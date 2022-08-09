import logging
from pathlib import PurePath, Path
from threading import main_thread
import numpy as np
from data_processor.dataset.ife_dataset import IFEDataset
from sklearn.metrics import confusion_matrix
from sklearn.preprocessing import LabelEncoder
import matplotlib.pyplot as plt
from timeit import default_timer as timer
import os
import json
from datetime import datetime

import pandas as pd
import xlsxwriter
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.styles import colors

import uuid
from tqdm import tqdm
import time

from utils.results_analysis import EpochDataProcessor
from vowel_synthesis.vowels_synthesizer import vowels_synthesizer
from data_processor.creators import get_SNR_dB
# from data_processor.labels_transformer import LabelsTransformer

def get_files(path, filename_mask):
    files = [file.name for file in Path(path).glob(filename_mask)]
    return files

def get_model_sheet_name(model_index, validation_signal_str):
    if validation_signal_str is None:
        model_sheet_name = f"M{model_index}"
    else:
        model_sheet_name = f"M{model_index}_{validation_signal_str}"
    return model_sheet_name

def get_params_from_model_sheet_name(model_sheet_name):
    # Valid names: "M{model_index}" or "M{model_index}_{validation_signal_str}"
    if model_sheet_name[0] != 'M':
        return None, None

    ind = 1
    while ind < len(model_sheet_name):
        if not ((model_sheet_name[ind] >= '0') and (model_sheet_name[ind] <= '9')):
            break
        ind += 1
    if ind == 1:
        return None, None
    
    model_index = int(model_sheet_name[1:ind])

    if ind == len(model_sheet_name):
        return model_index, None

    if model_sheet_name[ind] != '_':
        return None, None

    validation_signal_str = model_sheet_name[ind+1:]
    if len(validation_signal_str) >= 3:
        if validation_signal_str[-4:] == "Figs":
            return None, None

    return model_index, validation_signal_str

def find_inference_data(validation_dataset_subpath, training_setup_subpath):
    # find available inference data for given validation_dataset_subpath and training_setup_subpath
    epoch_infer_save_path = Path("./output_data/inference", validation_dataset_subpath, training_setup_subpath)

    epoch_strings = []
    epoch_indexes = []
    #names = get_files(data_path, "*.npz")
    names = get_files(epoch_infer_save_path, "infer_data_*.npz")
    for (idx, name) in enumerate(names):
        epoch_str = name[11:-4]
        epoch_strings.insert(idx, epoch_str)
        epoch_indexes.insert(idx, int(epoch_str[6:])) # TODO convert to number of None

    decorated = [(epoch_index, epoch_strings[idx], idx) for idx, epoch_index in enumerate(epoch_indexes)]
    decorated.sort()
    epoch_indexes = [epoch_index for epoch_index, epoch_string, idx in decorated]
    epoch_strings = [epoch_string for epoch_index, epoch_string, idx in decorated]


    return epoch_strings, epoch_indexes, epoch_infer_save_path

def draw_analyzis_results(epoch_dp, infer_results, setup_idx):
    processed_accuracy = epoch_dp.get_processed_accuracy()

    classes_labels = epoch_dp.classes_labels
    if classes_labels[()] is None:
        raise Exception(f"Could not find classes labels in {epochs_infer_save_path.as_posix()}; change folders or run inferer first")

    if classes_labels[0] == 0:
        classes_labels = classes_labels[1:]
    if classes_labels[2]-classes_labels[1] > classes_labels[1]-classes_labels[0]:
        # log classes
        factor = np.sqrt(classes_labels[2]/classes_labels[1])
        classes_edges = classes_labels / factor
        classes_edges = np.append(classes_edges, classes_labels[-1] * factor)
    else:
        # lin classes
        factor = (classes_labels[2] - classes_labels[1])/2
        classes_edges = classes_labels - factor
        classes_edges = np.append(classes_edges, classes_labels[-1] + factor)

    figs = {}
    figs[1] = plt.figure(f"{setup_idx}: model training accuracy")
    ax1 = figs[1].gca()
    # ax.plot(correct_table/total_table, 'r.')  # the same as model_accuracy_table
    ax1.plot(processed_accuracy["model_accuracy_table_all"], 'r-')  
    ax1.plot(processed_accuracy["d_correct_table"]/processed_accuracy["d_total_table"], 'b.')

    # # # Confusion Matrix
    # # # https://towardsdatascience.com/understanding-the-confusion-matrix-and-how-to-implement-it-in-python-319202e0fe4d
    # # Use classes_labels to convert frequencies to labels\
    # encoder = LabelEncoder()
    # encoder.fit(epoch_dp.classes_labels)
    # F0_ref_labels = encoder.transform(infer_results["F0_ref_labels"])
    # F0_est_labels = encoder.transform(infer_results["F0_est"])
    # # # labels_transformer = LabelsTransformer(classes_labels = classes_labels) 
    # infer_accuracy[model_epoch_idx] = (100.0 * (F0_ref_labels == F0_est_labels).sum())/len(F0_ref_labels)
    ax1.plot(np.arange(1,len(epoch_dp.exact_percentage)+1)*processed_accuracy["averaged_model_accuracy_table_size"]-1, epoch_dp.exact_percentage, 'ko')
    # ax1.plot(np.arange(1,model_epoch_idx+2)*processed_accuracy["averaged_model_accuracy_table_size"]-1, infer_accuracy[0:(infer_data_set_idx+1)], 'ko')
    ax1.plot(np.arange(1,len(epoch_dp.exact_percentage)+1)*processed_accuracy["averaged_model_accuracy_table_size"]-1, processed_accuracy["epochs_accuracy"], 'r*')

    figs[1].show()


    # CM = confusion_matrix(F0_ref_labels, F0_est_labels) # confusion_matrix(y_true, y_pred)
    # # https://stackoverflow.com/questions/35572000/how-can-i-plot-a-confusion-matrix
    # # https://matplotlib.org/stable/gallery/images_contours_and_fields/image_annotated_heatmap.html
    # # plt.figure(figsize = (10,7))
    # # sn.heatmap(CM, annot=True)
    # fig, ax = plt.subplots()
    # im = ax.imshow(CM, cmap="YlGn")
    # # Create colorbar
    # #, mappable=plt.cm.ScalarMappable(norm=None, cmap="YlGn")
    # cbar = ax.figure.colorbar(im, ax=ax) #, **cbar_kw)
    # # cbar.ax.set_ylabel(cbarlabel, rotation=-90, va="bottom")    
    # fig.tight_layout()
    # fig.show()


    # Map of logarithm of counts of estimation error values observed for
    # given reference 𝐹0 of synthetic validation signal for neural networks
    # with 351 classes and reference methods.
    F0_ref_resolution = 0.1 # in Hz
    F0_ref_q = np.round(infer_results["F0_ref"] / F0_ref_resolution)*F0_ref_resolution

    dF0 = infer_results["F0_ref"] - infer_results["F0_est"]
    dF0_ref_resolution = 0.1 # in Hz
    dF0_ref_q = np.round(dF0 / dF0_ref_resolution)*dF0_ref_resolution
    figs[2] = plt.figure(f"{setup_idx}: F0_ref error vs F0_ref")
    ax = figs[2].subplots()
    # ax.hist2d(F0_ref_q, dF0_ref_q, bins = [int((F0_ref_q.max()-F0_ref_q.min())/F0_ref_resolution), int((dF0_ref_q.max()-dF0_ref_q.min())/dF0_ref_resolution)])

    # F0_ref_edges = np.arange(50-F0_ref_resolution/2, 400+F0_ref_resolution/2,F0_ref_resolution)
    F0_ref_edges = classes_edges
    dF0_ref_edges = np.arange(-20-dF0_ref_resolution/2, 20+dF0_ref_resolution/2, dF0_ref_resolution)
    ax.hist2d(F0_ref_q, dF0_ref_q, bins = [F0_ref_edges, dF0_ref_edges], cmap=plt.get_cmap("gist_yarg"))
    #fig.colorbar()
    # https://matplotlib.org/stable/api/_as_gen/matplotlib.pyplot.pcolormesh.html
    figs[2].show()
    x = 1 

    tmp_figs = epoch_dp.plot_overall_data(setup_idx)
    for key in tmp_figs.keys():
        idx = len(figs)
        figs[idx+1] = tmp_figs[key]
    
    return figs

class ModelData2Excel:
    # =====================
    # Models sheet
    # best epoch index (based on training accuracy)
    BEST_TRAIN_EPOCH_INDEX_COL = -1
    # best training accuracy
    BEST_TRAIN_ACCU_COL = -1
    # best epoch index (based on inference accuracy)
    BEST_INFER_ACCU_EPOCH_INDEX_COL = -1
    # best inference accuracy
    BEST_INFER_ACCU_COL = -1
    # last epoch index 
    LAST_EPOCH_INDEX_COL = -1
    # last training accuracy
    LAST_TRAIN_ACCU_COL = -1
    # last inference accuracy
    LAST_INFER_ACCU_COL = -1

    # ++++++++++++++++++++++
    BEST_INFER_ACCU_OUTLIERS_COL = -1
    BEST_INFER_ACCU_MEAN_COL = -1
    BEST_INFER_ACCU_STD_COL = -1

    BEST_INFER_RMSE_EPOCH_INDEX_COL = -1
    BEST_INFER_RMSE_COL = -1
    BEST_INFER_RMSE_RELATIVE_EPOCH_INDEX_COL = -1
    BEST_INFER_RMSE_RELATIVE_COL = -1

    BEST_INFER_STD_EPOCH_INDEX_COL = -1
    BEST_INFER_STD_ACCU_COL = -1
    BEST_INFER_STD_OUTLIERS_COL = -1
    BEST_INFER_STD_MEAN_COL = -1
    BEST_INFER_STD_COL = -1

    BEST_INFER_STD_RELATIVE_EPOCH_INDEX_COL = -1
    BEST_INFER_STD_RELATIVE_ACCU_COL = -1
    BEST_INFER_STD_RELATIVE_OUTLIERS_COL = -1
    BEST_INFER_STD_RELATIVE_MEAN_COL = -1
    BEST_INFER_STD_RELATIVE_COL = -1

    # ++++++++++++++++++++++
    LAST_INFER_OUTLIERS_COL = -1
    LAST_INFER_MEAN_COL = -1
    LAST_INFER_STD_COL = -1
    LAST_INFER_MEAN_RELATIVE_COL = -1
    LAST_INFER_STD_RELATIVE_COL = -1
    # ++++++++++++++++++++++

    # =====================
    # Model parameters sheet
    UUID_ROW = 1
    NAME_ROW = 2
    MODEL_PATH_ROW = 3
    VALIDATION_MODE_ROW = 4
    # =====================
    EPOCH_NO_ROW = 6
    TRAIN_ACCU_ROW = 7
    TRAIN_SNR_ROW = 8
    INFER_ACCU_ROW = 9
    OUTLIERS_ROW = 10
    MEAN_ROW = 11
    STD_ROW = 12
    RMSE_ROW = 13
    R2_ROW = 14
    MEAN_RELATIVE_ROW = 15
    STD_RELATIVE_ROW = 16
    RMSE_RELATIVE_ROW = 17
    R2_RELATIVE_ROW = 18

    LABELS_ROW = 20

    def __init__(self, F0_relative_ref = 100.0) -> None:
        self.models_database_filename = None
        self.models_names = dict()
        self.model = None
        self.workbook4stage = None
        self.F0_relative_ref = F0_relative_ref

        self.set_of_colors = list()
        for color in colors.COLOR_INDEX:
            if color == '00FFFFFF':
                continue
            if color in self.set_of_colors:
                continue
            self.set_of_colors.append(color)

    # def get_color(self, index):
    #     if index == 0:
    #         raise Exception("color index 0 is not supported")
        
    #     # TODO https://stackoverflow.com/questions/42086276/get-default-line-colour-cycle
    #     # TODO https://matplotlib.org/stable/tutorials/colors/colormaps.html

    #     idx = ((index-1) % (len(colors.COLOR_INDEX)-1)) + 1
    #     if idx == 1:
    #         idx = 0
    #     return colors.COLOR_INDEX[idx]

    def get_color(self, index):
        if index == 0:
            raise Exception("color index 0 is not supported")

        idx = ((index-1) % len(self.set_of_colors))
        return self.set_of_colors[idx]

    def load_database(self):
        if self.workbook4stage is None:
            # load excel workbook
            if Path(self.models_database_filename).is_file():
                # models_data = pd.read_excel(models_database_filename, sheet_name='Models')
                self.workbook4stage = openpyxl.load_workbook(self.models_database_filename)

    def stage_model_parameters(self, epoch_index, validation_signal_str = None, training_accuracy=None, inference_accuracy=None,
            outliers_percentage=None, error_without_outliers_mean=None, error_without_outliers_std=None,
            error_without_outliers_mean_relative = None, error_without_outliers_std_relative=None, trained_SNR_dB=None):
        # collects model parameters and updates excel workbook based on self.models_database_filename
        # NOTE when all data are stored call commit_model_parameters (or find_model_in_database) to save them
        if self.models_database_filename is None:
            raise Exception("find_model_in_database has to be called first")

        # at first call load excel workbook
        self.load_database()

        # make changes in model's sheet of loaded workbook
        model_index = self.model["Index"]

        model_sheet_name = get_model_sheet_name(model_index, validation_signal_str)
            
        model_data_worksheet = self.workbook4stage[model_sheet_name]

        model_data_worksheet.cell(self.EPOCH_NO_ROW, 2+epoch_index+1, value=epoch_index) 
        if trained_SNR_dB is not None:
            model_data_worksheet.cell(self.TRAIN_SNR_ROW, 2+epoch_index+1, value=trained_SNR_dB.item())  # extract np scalar
        if training_accuracy is not None:
            model_data_worksheet.cell(self.TRAIN_ACCU_ROW, 2+epoch_index+1, value=training_accuracy) 
        if inference_accuracy is not None:
            model_data_worksheet.cell(self.INFER_ACCU_ROW, 2+epoch_index+1, value=inference_accuracy) 
        if outliers_percentage is not None:
            model_data_worksheet.cell(self.OUTLIERS_ROW, 2+epoch_index+1, value=outliers_percentage) 
        if error_without_outliers_mean is not None:
            model_data_worksheet.cell(self.MEAN_ROW, 2+epoch_index+1, value=error_without_outliers_mean) 
        if error_without_outliers_std is not None:
            model_data_worksheet.cell(self.STD_ROW, 2+epoch_index+1, value=error_without_outliers_std) 
            
            RMSE = np.sqrt(np.power(error_without_outliers_std,2) + np.power(error_without_outliers_mean,2))
            model_data_worksheet.cell(self.RMSE_ROW, 2+epoch_index+1, value=RMSE) 
            
            # read RMSE for epoch_index == 0
            RMSE_0 = model_data_worksheet.cell(self.RMSE_ROW, 2+0+1).value 
            if RMSE_0 is None:
                RMSE_0 = RMSE
            # calculate and save R^2 
            R2 = (np.power(RMSE_0,2) - np.power(RMSE,2))/(np.power(RMSE_0,2))
            model_data_worksheet.cell(self.R2_ROW, 2+epoch_index+1, value=R2) 

        if error_without_outliers_mean_relative is not None:
            model_data_worksheet.cell(self.MEAN_RELATIVE_ROW, 2+epoch_index+1, value=error_without_outliers_mean_relative * self.F0_relative_ref) 
        if error_without_outliers_std_relative is not None:
            model_data_worksheet.cell(self.STD_RELATIVE_ROW, 2+epoch_index+1, value=error_without_outliers_std_relative * self.F0_relative_ref) 
            
            RMSE_rel = np.sqrt(np.power(error_without_outliers_std_relative * self.F0_relative_ref, 2) 
                    + np.power(error_without_outliers_mean_relative * self.F0_relative_ref, 2))
            model_data_worksheet.cell(self.RMSE_RELATIVE_ROW, 2+epoch_index+1, value=RMSE_rel) 
            
            # read RMSE for epoch_index == 0
            RMSE_0_rel = model_data_worksheet.cell(self.RMSE_RELATIVE_ROW, 2+0+1).value 
            if RMSE_0_rel is None:
                RMSE_0_rel = RMSE_rel
            # calculate and save R^2 
            R2_rel = (np.power(RMSE_0_rel,2) - np.power(RMSE_rel,2))/(np.power(RMSE_0_rel,2))
            model_data_worksheet.cell(self.R2_RELATIVE_ROW, 2+epoch_index+1, value=R2_rel) 

        # self.update_best_model_results(model_data_worksheet, model_index)

    def update_best_results(self):

        max_no_of_epochs = 0
        print("Updating best results...")
        with tqdm(total=len(self.workbook4stage.sheetnames), dynamic_ncols=True) as progressbar:
            for model_sheet_name in self.workbook4stage.sheetnames:
                #  get model_index and validation_signal_str
                model_index, validation_signal_str = get_params_from_model_sheet_name(model_sheet_name)
                if model_index is not None:
                    model_sheet_name = get_model_sheet_name(model_index, validation_signal_str)
                        
                    model_data_worksheet = self.workbook4stage[model_sheet_name]
                    no_of_epochs = self.update_best_model_results(model_data_worksheet, model_index)
                    if max_no_of_epochs < no_of_epochs:
                        max_no_of_epochs = no_of_epochs
                progressbar.update(1)
        return max_no_of_epochs

    def update_best_model_results(self, model_data_worksheet, model_index):
        # TODO read stored parameters of the current model
        # https://www.tutorialsteacher.com/python/python-generator
        model_data_worksheet_rows_list = list(model_data_worksheet.rows)

        epochs_row = model_data_worksheet_rows_list[self.EPOCH_NO_ROW-1]
        epochs_indexes = [idx-2 for idx, epoch_cell in enumerate(epochs_row) if idx > 1 and epoch_cell.value is not None]
        # self.no_of_epochs = len(epochs_indexes)
        if len(epochs_indexes) == 0:
            self.no_of_epochs = 0
            print("")
            print(f"Model {model_index} has no epoch data")
            return 0
        else:
            self.no_of_epochs = max(epochs_indexes)
        last_training_accuracy = model_data_worksheet.cell(self.TRAIN_ACCU_ROW, 2+self.no_of_epochs).value 
        last_inference_accuracy = model_data_worksheet.cell(self.INFER_ACCU_ROW, 2+self.no_of_epochs).value 

        train_accu_row = model_data_worksheet_rows_list[self.TRAIN_ACCU_ROW-1]
        train_accuracies = [train_accu_cell.value for idx, train_accu_cell in enumerate(train_accu_row) if idx > 1 and train_accu_cell.value is not None]
        max_train_accuracy_idx = train_accuracies.index(max(train_accuracies))
        # model_data_worksheet.row_dimensions[self.TRAIN_ACCU_ROW]
        for idx in range(self.no_of_epochs):
            model_data_worksheet.cell(self.TRAIN_ACCU_ROW, 2+idx+1).font = Font(bold=False) 
        model_data_worksheet.cell(self.TRAIN_ACCU_ROW, 2+max_train_accuracy_idx+1).font = Font(bold=True) 

        infer_accu_row = model_data_worksheet_rows_list[self.INFER_ACCU_ROW-1]
        infer_accuracies = [infer_accu_cell.value for idx, infer_accu_cell in enumerate(infer_accu_row) if idx > 1 and infer_accu_cell.value is not None]
        max_infer_accuracy_idx = infer_accuracies.index(max(infer_accuracies))
        for idx in range(self.no_of_epochs):
            model_data_worksheet.cell(self.INFER_ACCU_ROW, 2+idx+1).font = Font(bold=False) 
        model_data_worksheet.cell(self.INFER_ACCU_ROW, 2+max_infer_accuracy_idx+1).font = Font(bold=True) 

        infer_outliers_row = model_data_worksheet_rows_list[self.OUTLIERS_ROW-1]
        infer_outliers = [infer_outliers_cell.value for idx, infer_outliers_cell in enumerate(infer_outliers_row) if idx > 1 and infer_outliers_cell.value is not None]
        min_infer_outliers_idx = infer_outliers.index(min(infer_outliers))
        for idx in range(self.no_of_epochs):
            model_data_worksheet.cell(self.OUTLIERS_ROW, 2+idx+1).font = Font(bold=False) 
        model_data_worksheet.cell(self.OUTLIERS_ROW, 2+min_infer_outliers_idx+1).font = Font(bold=True) 

        infer_mean_row = model_data_worksheet_rows_list[self.MEAN_ROW-1]
        infer_means = [infer_mean_cell.value for idx, infer_mean_cell in enumerate(infer_mean_row) if idx > 1 and infer_mean_cell.value is not None]
        best_infer_mean_idx = np.argmin(np.abs(infer_means))
        for idx in range(self.no_of_epochs):
            model_data_worksheet.cell(self.MEAN_ROW, 2+idx+1).font = Font(bold=False) 
        model_data_worksheet.cell(self.MEAN_ROW, 2+best_infer_mean_idx+1).font = Font(bold=True) 

        infer_std_row = model_data_worksheet_rows_list[self.STD_ROW-1]
        infer_stds = [infer_std_cell.value for idx, infer_std_cell in enumerate(infer_std_row) if idx > 1 and infer_std_cell.value is not None]
        min_infer_std_idx = infer_stds.index(min(infer_stds))
        for idx in range(self.no_of_epochs):
            model_data_worksheet.cell(self.STD_ROW, 2+idx+1).font = Font(bold=False) 
        model_data_worksheet.cell(self.STD_ROW, 2+min_infer_std_idx+1).font = Font(bold=True) 

        infer_RMSE_row = model_data_worksheet_rows_list[self.RMSE_ROW-1]
        infer_RMSEs = [infer_RMSE_cell.value for idx, infer_RMSE_cell in enumerate(infer_RMSE_row) if idx > 1 and infer_RMSE_cell.value is not None]
        min_infer_RMSE_idx = infer_RMSEs.index(min(infer_RMSEs))
        for idx in range(self.no_of_epochs):
            model_data_worksheet.cell(self.RMSE_ROW, 2+idx+1).font = Font(bold=False) 
        model_data_worksheet.cell(self.RMSE_ROW, 2+min_infer_RMSE_idx+1).font = Font(bold=True) 

        infer_mean_rel_row = model_data_worksheet_rows_list[self.MEAN_RELATIVE_ROW-1]
        infer_means_rel = [infer_mean_rel_cell.value for idx, infer_mean_rel_cell in enumerate(infer_mean_rel_row) if idx > 1 and infer_mean_rel_cell.value is not None]
        if len(infer_means_rel) > 0:
            best_infer_mean_rel_idx = np.argmin(np.abs(infer_means_rel))
            for idx in range(self.no_of_epochs):
                model_data_worksheet.cell(self.MEAN_RELATIVE_ROW, 2+idx+1).font = Font(bold=False) 
            model_data_worksheet.cell(self.MEAN_RELATIVE_ROW, 2+best_infer_mean_rel_idx+1).font = Font(bold=True) 

        infer_std_rel_row = model_data_worksheet_rows_list[self.STD_RELATIVE_ROW-1]
        infer_stds_rel = [infer_std_rel_cell.value for idx, infer_std_rel_cell in enumerate(infer_std_rel_row) if idx > 1 and infer_std_rel_cell.value is not None]
        if len(infer_stds_rel) > 0:
            min_infer_std_rel_idx = infer_stds_rel.index(min(infer_stds_rel))
            for idx in range(self.no_of_epochs):
                model_data_worksheet.cell(self.STD_RELATIVE_ROW, 2+idx+1).font = Font(bold=False) 
            model_data_worksheet.cell(self.STD_RELATIVE_ROW, 2+min_infer_std_rel_idx+1).font = Font(bold=True) 

        infer_RMSE_rel_row = model_data_worksheet_rows_list[self.RMSE_RELATIVE_ROW-1]
        infer_RMSEs_rel = [infer_RMSE_rel_cell.value for idx, infer_RMSE_rel_cell in enumerate(infer_RMSE_rel_row) if idx > 1 and infer_RMSE_rel_cell.value is not None]
        min_infer_RMSE_rel_idx = infer_RMSEs_rel.index(min(infer_RMSEs_rel))
        for idx in range(self.no_of_epochs):
            model_data_worksheet.cell(self.RMSE_RELATIVE_ROW, 2+idx+1).font = Font(bold=False) 
        model_data_worksheet.cell(self.RMSE_RELATIVE_ROW, 2+min_infer_RMSE_rel_idx+1).font = Font(bold=True) 
        

        # TODO update best epoch parameter in main sheet of the loaded workbook
        models_worksheet = self.workbook4stage["Models"]
        # best epoch index (based on training accuracy)
        models_worksheet.cell(1+model_index, self.BEST_TRAIN_EPOCH_INDEX_COL, value=max_train_accuracy_idx) 
        # best training accuracy
        models_worksheet.cell(1+model_index, self.BEST_TRAIN_ACCU_COL, value=train_accuracies[max_train_accuracy_idx]) 
        
        # best epoch index (based on inference accuracy)
        models_worksheet.cell(1+model_index, self.BEST_INFER_ACCU_EPOCH_INDEX_COL, value=max_infer_accuracy_idx) 
        # best inference accuracy
        models_worksheet.cell(1+model_index, self.BEST_INFER_ACCU_COL, value=infer_accuracies[max_infer_accuracy_idx]) 
        # best infer accu outliers
        models_worksheet.cell(1+model_index, self.BEST_INFER_ACCU_OUTLIERS_COL, value=infer_outliers[max_infer_accuracy_idx]) 
        # best infer accu mean
        models_worksheet.cell(1+model_index, self.BEST_INFER_ACCU_MEAN_COL, value=infer_means[max_infer_accuracy_idx]) 
        # best infer accu std
        models_worksheet.cell(1+model_index, self.BEST_INFER_ACCU_STD_COL, value=infer_stds[max_infer_accuracy_idx]) 

        # ++++++++++++++++++++++
        # best epoch index (based on RMSE)
        models_worksheet.cell(1+model_index, self.BEST_INFER_RMSE_EPOCH_INDEX_COL, value=min_infer_RMSE_idx) 
        # best infer RMSE
        models_worksheet.cell(1+model_index, self.BEST_INFER_RMSE_COL, value=infer_RMSEs[min_infer_RMSE_idx]) 
        # best epoch index (based on relative RMSE)
        models_worksheet.cell(1+model_index, self.BEST_INFER_RMSE_RELATIVE_EPOCH_INDEX_COL, value=min_infer_RMSE_rel_idx) 
        # best infer relative RMSE
        models_worksheet.cell(1+model_index, self.BEST_INFER_RMSE_RELATIVE_COL, value=infer_RMSEs_rel[min_infer_RMSE_rel_idx]) 

        # ++++++++++++++++++++++
        # best epoch index (based on std)
        models_worksheet.cell(1+model_index, self.BEST_INFER_STD_EPOCH_INDEX_COL, value=min_infer_std_idx) 
        # best infer std accu
        models_worksheet.cell(1+model_index, self.BEST_INFER_STD_ACCU_COL, value=infer_accuracies[min_infer_std_idx]) 
        # best infer std outliers
        models_worksheet.cell(1+model_index, self.BEST_INFER_STD_OUTLIERS_COL, value=infer_outliers[min_infer_std_idx]) 
        # best infer std mean
        models_worksheet.cell(1+model_index, self.BEST_INFER_STD_MEAN_COL, value=infer_means[min_infer_std_idx]) 
        # best infer std std
        models_worksheet.cell(1+model_index, self.BEST_INFER_STD_COL, value=infer_stds[min_infer_std_idx]) 
        # ++++++++++++++++++++++

# TODO min_infer_std_rel_idx might not be defined !!!
        # best epoch index (based on std_rel)
        models_worksheet.cell(1+model_index, self.BEST_INFER_STD_RELATIVE_EPOCH_INDEX_COL, value=min_infer_std_rel_idx) 
        # best infer std accu
        models_worksheet.cell(1+model_index, self.BEST_INFER_STD_RELATIVE_ACCU_COL, value=infer_accuracies[min_infer_std_rel_idx]) 
        # best infer std outliers
        models_worksheet.cell(1+model_index, self.BEST_INFER_STD_RELATIVE_OUTLIERS_COL, value=infer_outliers[min_infer_std_rel_idx]) 
        # best infer std mean
        models_worksheet.cell(1+model_index, self.BEST_INFER_STD_RELATIVE_MEAN_COL, value=infer_means_rel[min_infer_std_rel_idx]) 
        # best infer std std
        models_worksheet.cell(1+model_index, self.BEST_INFER_STD_RELATIVE_COL, value=infer_stds_rel[min_infer_std_rel_idx]) 
        # ++++++++++++++++++++++

        # last epoch index 
        models_worksheet.cell(1+model_index, self.LAST_EPOCH_INDEX_COL, value=self.no_of_epochs-1) 
        # last training accuracy
        models_worksheet.cell(1+model_index, self.LAST_TRAIN_ACCU_COL, value=last_training_accuracy) 
        # last inference accuracy
        models_worksheet.cell(1+model_index, self.LAST_INFER_ACCU_COL, value=last_inference_accuracy) 
        # last infer outliers
        models_worksheet.cell(1+model_index, self.LAST_INFER_OUTLIERS_COL, value=infer_outliers[-1]) 
        # last infer mean
        models_worksheet.cell(1+model_index, self.LAST_INFER_MEAN_COL, value=infer_means[-1]) 
        # last infer std
        models_worksheet.cell(1+model_index, self.LAST_INFER_STD_COL, value=infer_stds[-1]) 
        # last infer mean relative
        models_worksheet.cell(1+model_index, self.LAST_INFER_MEAN_RELATIVE_COL, value=infer_means_rel[-1]) 
        # last infer std relative
        models_worksheet.cell(1+model_index, self.LAST_INFER_STD_RELATIVE_COL, value=infer_stds_rel[-1]) 

        return self.no_of_epochs

    def add_model_classes_labels(self, classes_labels, validation_signal_str):
        # at first call load excel workbook
        self.load_database()

        if self.model["use_F0_too_low_class"] == False and classes_labels[0] == 0:
            classes_labels = classes_labels[1:]
        self.classes_labels = classes_labels

        # make changes in model's sheet of loaded workbook
        model_index = self.model["Index"]

        model_sheet_name = get_model_sheet_name(model_index, validation_signal_str)
        model_data_worksheet = self.workbook4stage[model_sheet_name]

        step = 6
        for idx, label in enumerate(classes_labels):
            model_data_worksheet.cell(self.LABELS_ROW+1+idx*step, 1, value=label)

    def add_parameters_per_class(self, validation_signal_str, no_of_data_per_class_all, no_of_outliers_per_class_all, 
                                 est_mean_error_per_class_all, est_std_error_per_class_all):
        # at first call load excel workbook
        self.load_database()

        # make changes in model's sheet of loaded workbook
        model_index = self.model["Index"]

        model_sheet_name = get_model_sheet_name(model_index, validation_signal_str)
        model_data_worksheet = self.workbook4stage[model_sheet_name]

        step = 6
        for epoch_idx, no_of_data_per_class in enumerate(no_of_data_per_class_all):
            min_idx = np.argmin(no_of_data_per_class)
            max_idx = np.argmax(no_of_data_per_class)
            for idx, no_of_data in enumerate(no_of_data_per_class):
                if epoch_idx == 0:
                    model_data_worksheet.cell(self.LABELS_ROW+1+idx*step, 2, "no_of_data").font = Font(bold=True)
                font = Font(bold=False) 
                if idx == min_idx:
                    font = Font(bold=True,color='00FF0000') # red
                if idx == max_idx:
                    font = Font(bold=True,color='00008000') # green
                model_data_worksheet.cell(self.LABELS_ROW+1+idx*step, 2+1+epoch_idx, value=no_of_data).font = font
        for epoch_idx, no_of_outliers_per_class in enumerate(no_of_outliers_per_class_all):
            min_idx = np.argmin(no_of_outliers_per_class)
            max_idx = np.argmax(no_of_outliers_per_class)
            for idx, no_of_outliers in enumerate(no_of_outliers_per_class):
                if epoch_idx == 0:
                    model_data_worksheet.cell(self.LABELS_ROW+1+idx*step + 1, 2, "no_of_outliers [%]").font = Font(bold=True)
                    if len("no_of_outliers [%]") + 1 > model_data_worksheet.column_dimensions[get_column_letter(2)].width:
                        model_data_worksheet.column_dimensions[get_column_letter(2)].width = len("no_of_outliers [%]") + 1
                font = Font(bold=False) 
                if idx == min_idx:
                    font = Font(bold=True,color='00FF0000') # red
                if idx == max_idx:
                    font = Font(bold=True,color='00008000') # green
                model_data_worksheet.cell(self.LABELS_ROW+1+idx*step + 1, 2+1+epoch_idx, value=float(100*no_of_outliers)/no_of_data).font = font 
        for epoch_idx, est_mean_error_per_class in enumerate(est_mean_error_per_class_all):
            min_idx = np.argmin(np.abs(est_mean_error_per_class))
            min_rel_idx = np.argmin(np.abs(est_mean_error_per_class)/self.classes_labels * self.F0_relative_ref)
            max_idx = np.argmax(np.abs(est_mean_error_per_class))
            max_rel_idx = np.argmax(np.abs(est_mean_error_per_class)/self.classes_labels * self.F0_relative_ref)
            for idx, est_mean_error in enumerate(est_mean_error_per_class):
                if epoch_idx == 0:
                    model_data_worksheet.cell(self.LABELS_ROW+1+idx*step + 2, 2, "est_mean_error").font = Font(bold=True)
                    model_data_worksheet.cell(self.LABELS_ROW+1+idx*step + 4, 2, "est_mean_relative").font = Font(bold=True)
                font = Font(bold=False) 
                if idx == min_idx:
                    font = Font(bold=True,color='00FF0000') # red
                if idx == max_idx:
                    font = Font(bold=True,color='00008000') # green
                model_data_worksheet.cell(self.LABELS_ROW+1+idx*step + 2, 2+1+epoch_idx, value=est_mean_error).font = font  
                # TODO: use formula                
                font = Font(bold=False) 
                if idx == min_rel_idx:
                    font = Font(bold=True,color='00FF0000') # red
                if idx == max_rel_idx:
                    font = Font(bold=True,color='00008000') # green
                model_data_worksheet.cell(self.LABELS_ROW+1+idx*step + 4, 2+1+epoch_idx, value=est_mean_error/self.classes_labels[idx] * self.F0_relative_ref).font = font  
        for epoch_idx, est_std_error_per_class in enumerate(est_std_error_per_class_all):
            min_idx = np.argmin(np.abs(est_std_error_per_class))
            min_rel_idx = np.argmin(np.abs(est_std_error_per_class)/self.classes_labels * self.F0_relative_ref)
            max_idx = np.argmax(np.abs(est_std_error_per_class))
            max_rel_idx = np.argmax(np.abs(est_std_error_per_class)/self.classes_labels * self.F0_relative_ref)
            for idx, est_std_error in enumerate(est_std_error_per_class):
                if epoch_idx == 0:
                    model_data_worksheet.cell(self.LABELS_ROW+1+idx*step + 3, 2, "est_std_error").font = Font(bold=True)
                    model_data_worksheet.cell(self.LABELS_ROW+1+idx*step + 5, 2, "est_std_relative").font = Font(bold=True)  
                font = Font(bold=False) 
                if idx == min_idx:
                    font = Font(bold=True,color='00FF0000') # red
                if idx == max_idx:
                    font = Font(bold=True,color='00008000') # green
                model_data_worksheet.cell(self.LABELS_ROW+1+idx*step + 3, 2+1+epoch_idx, value=est_std_error).font = font  
                # TODO: use formula
                font = Font(bold=False) 
                if idx == min_rel_idx:
                    font = Font(bold=True,color='00FF0000') # red
                if idx == max_rel_idx:
                    font = Font(bold=True,color='00008000') # green
                model_data_worksheet.cell(self.LABELS_ROW+1+idx*step + 5, 2+1+epoch_idx, value=est_std_error/self.classes_labels[idx] * self.F0_relative_ref).font = font  
        

    def commit_model_parameters(self, validation_signal_str):
        # save all staged model parametrs to Excel file
        if self.workbook4stage is not None:
            max_no_of_epochs = self.update_best_results()
            self.add_models_chart(validation_signal_str, max_no_of_epochs)
            self.workbook4stage.save(self.models_database_filename)
            self.workbook4stage = None

    def find_model_in_database(self, model_params, models_database_filename, validation_signal_str, validation_SNR_str):
        #  checks if models is in database if no, adds it to it
        #  additionally sets the self.model variable to current model data with updated UUID, index and name

        # If model parameters are not committed then do this now
        self.commit_model_parameters(validation_signal_str)

        self.models_database_filename = models_database_filename

        model_ID = None

        workbook = None
        models_worksheet = None
        if Path(models_database_filename).is_file():
            # models_data = pd.read_excel(models_database_filename, sheet_name='Models')
            workbook = openpyxl.load_workbook(models_database_filename)
            models_worksheet = workbook["Models"]
        
        # model = pd.DataFrame({
        #     "UUID": uuid.uuid4(),
        #     "name": "???",
        #     "model_type": models_data["model_config"]["type"],
        #     "model_params": models_data["model_config"]["params"],
        #     "no_of_inputs": models_data["model_config"]["no_of_inputs"],
        #     "no_of_classes": models_data["model_config"]["no_of_classes"],
        #     "classes_mode": models_data["labels_transformer_options"]["mode"],
        #     "F0_ref_min": models_data["labels_transformer_options"]["F0_ref_min"],
        #     "F0_ref_max": models_data["labels_transformer_options"]["F0_ref_max"],
        #     "use_F0_too_low_class": models_data["labels_transformer_options"]["use_F0_too_low_class"],
        #     "learning_rate": models_data["checkpoint_trainer_config"]["learning_rate"],
        #     "batch_size": models_data["checkpoint_trainer_config"]["batch_size"],
        #     "model_path": models_data["model_path"]
        # }

        training_mode = model_params["training_data_options"]["mode"]
        if training_mode == "file":
            training_SNR_dB = "-"
        else:
            SNR_tab = model_params["training_data_options"]["SNR_dB"]
            training_SNR_dB = f"{SNR_tab[0]:.0f}-{SNR_tab[1]:.0f}"

        validation_mode = model_params["validation_data_options"]["mode"]
        if validation_mode == "file":
            SNR_dB = "-"
            no_of_segments = -1
            F0_ref_filename = model_params["validation_data_options"]["F0_ref_filename"]
        else:
            SNR_tab = model_params["validation_data_options"]["SNR_dB"]
            SNR_dB = f"{SNR_tab[0]:.0f}-{SNR_tab[1]:.0f}"
            no_of_segments = model_params["validation_data_options"]["no_of_segments"]
            F0_ref_filename = "-"

        self.model = {
            "Index": -1,
            "UUID": "", # uuid.uuid4(),
            "name": "", # TODO add option to define it in configuration
            "model_type": model_params["model_config"]["type"],
            "model_params": json.dumps(model_params["model_config"]["parameters"]),
            "no_of_inputs": model_params["model_config"]["no_of_inputs"],
            "no_of_classes": model_params["model_config"]["no_of_classes"],
            "classes_mode": model_params["labels_transformer_options"]["mode"],
            "F0_ref_min": model_params["labels_transformer_options"]["F0_ref_min"],
            "F0_ref_max": model_params["labels_transformer_options"]["F0_ref_max"],
            "use_F0_too_low_class": model_params["labels_transformer_options"]["use_F0_too_low_class"],
            "power normal.": model_params["dataset_options"]["FB_data power normalization mode"],
            "freq normal.": model_params["dataset_options"]["FB_data frequency normalization mode"],
            "learning_rate": model_params["checkpoint_trainer_config"]["learning_rate"],
            "FB_to_CMPO_smoothing_factor": model_params["dataset_options"]["AFE"]["FB_to_CMPO_smoothing_factor"],
            "batch_size": model_params["checkpoint_trainer_config"]["batch_size"],
            "best train epoch idx": -1,
            "best train accu": -1,
            "best infer accu epoch idx": -1,
            "best infer accu": -1,
            
            # ++++++++++++++++++++++
            "best infer accu outliers": -1,
            "best infer accu mean": -1,
            "best infer accu std": -1,

            "best infer RMSE epoch idx": -1,
            "best infer RMSE": -1,
            "best infer rel. RMSE epoch idx": -1,
            "best infer rel. RMSE": -1,

            "best infer std epoch idx": -1,
            "best infer std accu": -1,
            
            "best infer std outliers": -1,
            "best infer std mean": -1,
            "best infer std": -1,
            # ++++++++++++++++++++++

            "best infer std_rel epoch idx": -1,
            "best infer std_rel accu": -1,
            
            "best infer std_rel outliers": -1,
            "best infer std_rel mean": -1,
            "best infer std_rel": -1,
            # ++++++++++++++++++++++

            "last epoch idx": -1,
            "last train accu": -1,
            "last infer accu": -1,

            # ++++++++++++++++++++++
            "last infer outliers": -1,
            "last infer mean": -1,
            "last infer std": -1,
            "last infer mean rel.": -1,
            "last infer std rel.": -1,
            # ++++++++++++++++++++++

            "training_mode": training_mode,
            "training_SNR_dB": training_SNR_dB,
            "training_reload": model_params["training_data_options"]["reload"],

            "validation_mode": validation_mode,
            "no_of_segments": no_of_segments,
            "SNR_dB": SNR_dB,
            "F0_ref_filename": F0_ref_filename,
            "validation_SNR_str": validation_SNR_str,
            "model_path": model_params["model_path"],
        }

        # comparison of medels by their parameters
        # \TODO use also FB mode and smoothing modes for FB and CMPO
        model_comparison_keys = ["model_type", "model_params", "no_of_inputs", "no_of_classes", "classes_mode", "F0_ref_min", "F0_ref_max",
            "power normal.", "freq normal.",
            "use_F0_too_low_class", "learning_rate", "batch_size",
            "training_mode", "training_SNR_dB", "training_reload", 
            "validation_mode", "no_of_segments", "SNR_dB", "FB_to_CMPO_smoothing_factor", 
            "F0_ref_filename", "validation_SNR_str"] # ? "model_path"

        list_of_keys = list(self.model.keys())

        do_update_constants = ( (self.BEST_TRAIN_EPOCH_INDEX_COL == -1) or 
                                (self.BEST_TRAIN_ACCU_COL == -1) or
                                (self.BEST_INFER_ACCU_EPOCH_INDEX_COL == -1) or
                                (self.BEST_INFER_ACCU_COL == -1) or
                                (self.LAST_EPOCH_INDEX_COL == -1) or
                                (self.LAST_TRAIN_ACCU_COL == -1) or
                                (self.LAST_INFER_ACCU_COL == -1) or
                                # ++++++++++++++++++++++
                                (self.BEST_INFER_ACCU_OUTLIERS_COL == -1) or
                                (self.BEST_INFER_ACCU_MEAN_COL == -1) or
                                (self.BEST_INFER_ACCU_STD_COL == -1) or

                                (self.BEST_INFER_RMSE_EPOCH_INDEX_COL == -1) or
                                (self.BEST_INFER_RMSE_COL == -1) or
                                (self.BEST_INFER_RMSE_RELATIVE_EPOCH_INDEX_COL == -1) or
                                (self.BEST_INFER_RMSE_RELATIVE_COL == -1) or

                                (self.BEST_INFER_STD_EPOCH_INDEX_COL == -1) or
                                (self.BEST_INFER_STD_ACCU_COL == -1) or
                                (self.BEST_INFER_STD_OUTLIERS_COL == -1) or
                                (self.BEST_INFER_STD_MEAN_COL == -1) or
                                (self.BEST_INFER_STD_COL == -1) or

                                (self.BEST_INFER_STD_RELATIVE_EPOCH_INDEX_COL == -1) or
                                (self.BEST_INFER_STD_RELATIVE_ACCU_COL == -1) or
                                (self.BEST_INFER_STD_RELATIVE_OUTLIERS_COL == -1) or
                                (self.BEST_INFER_STD_RELATIVE_MEAN_COL == -1) or
                                (self.BEST_INFER_STD_RELATIVE_COL == -1) or

                                # ++++++++++++++++++++++
                                (self.LAST_INFER_OUTLIERS_COL == -1) or
                                (self.LAST_INFER_MEAN_COL == -1) or
                                (self.LAST_INFER_STD_COL == -1) or
                                (self.LAST_INFER_MEAN_RELATIVE_COL == -1) or
                                (self.LAST_INFER_STD_RELATIVE_COL == -1))
                                # ++++++++++++++++++++++
        if  do_update_constants:
            for key_idx, key_name in enumerate(list_of_keys):
                if key_name == "best train epoch idx":
                    self.BEST_TRAIN_EPOCH_INDEX_COL = key_idx+1
                if key_name == "best train accu":
                    self.BEST_TRAIN_ACCU_COL = key_idx+1
                if key_name == "best infer accu epoch idx":
                   self.BEST_INFER_ACCU_EPOCH_INDEX_COL = key_idx+1
                if key_name == "best infer accu":
                    self.BEST_INFER_ACCU_COL = key_idx+1
                if key_name == "last epoch idx":
                    self.LAST_EPOCH_INDEX_COL = key_idx+1
                if key_name == "last train accu": 
                    self.LAST_TRAIN_ACCU_COL = key_idx+1
                if key_name == "last infer accu": 
                    self.LAST_INFER_ACCU_COL = key_idx+1
                # ++++++++++++++++++++++
                if key_name == "best infer accu outliers": 
                    self.BEST_INFER_ACCU_OUTLIERS_COL = key_idx+1
                if key_name == "best infer accu mean": 
                    self.BEST_INFER_ACCU_MEAN_COL = key_idx+1
                if key_name == "best infer accu std": 
                    self.BEST_INFER_ACCU_STD_COL = key_idx+1


                if key_name == "best infer RMSE epoch idx": 
                    self.BEST_INFER_RMSE_EPOCH_INDEX_COL = key_idx+1
                if key_name == "best infer RMSE": 
                    self.BEST_INFER_RMSE_COL = key_idx+1
                if key_name == "best infer rel. RMSE epoch idx": 
                    self.BEST_INFER_RMSE_RELATIVE_EPOCH_INDEX_COL = key_idx+1
                if key_name == "best infer rel. RMSE": 
                    self.BEST_INFER_RMSE_RELATIVE_COL = key_idx+1

                if key_name == "best infer std epoch idx": 
                    self.BEST_INFER_STD_EPOCH_INDEX_COL = key_idx+1
                if key_name == "best infer std accu": 
                    self.BEST_INFER_STD_ACCU_COL = key_idx+1
                if key_name == "best infer std outliers": 
                    self.BEST_INFER_STD_OUTLIERS_COL = key_idx+1
                if key_name == "best infer std mean": 
                    self.BEST_INFER_STD_MEAN_COL = key_idx+1
                if key_name == "best infer std": 
                    self.BEST_INFER_STD_COL = key_idx+1

                # ++++++++++++++++++++++
                if key_name == "best infer std_rel epoch idx": 
                    self.BEST_INFER_STD_RELATIVE_EPOCH_INDEX_COL = key_idx+1
                if key_name == "best infer std_rel accu": 
                    self.BEST_INFER_STD_RELATIVE_ACCU_COL = key_idx+1
                if key_name == "best infer std_rel outliers": 
                    self.BEST_INFER_STD_RELATIVE_OUTLIERS_COL = key_idx+1
                if key_name == "best infer std_rel mean": 
                    self.BEST_INFER_STD_RELATIVE_MEAN_COL = key_idx+1
                if key_name == "best infer std_rel": 
                    self.BEST_INFER_STD_RELATIVE_COL = key_idx+1

                # ++++++++++++++++++++++
                if key_name == "last infer outliers": 
                    self.LAST_INFER_OUTLIERS_COL = key_idx+1
                if key_name == "last infer mean": 
                    self.LAST_INFER_MEAN_COL = key_idx+1
                if key_name == "last infer std": 
                    self.LAST_INFER_STD_COL = key_idx+1
                if key_name == "last infer mean rel.": 
                    self.LAST_INFER_MEAN_RELATIVE_COL = key_idx+1
                if key_name == "last infer std rel.": 
                    self.LAST_INFER_STD_RELATIVE_COL = key_idx+1
                # ++++++++++++++++++++++

        if models_worksheet is not None:
            # find model in models_data
            no_of_models = len(models_worksheet['A']) - 1

            # model_type_col = list_of_keys.index("model_type") + 1
            UUID_col = list_of_keys.index("UUID") + 1
            name_col = list_of_keys.index("name") + 1
            for row in range(2,2+no_of_models):
                ok_counter = 0

                model_Index = models_worksheet.cell(row=row, column=1).value
                model_name = models_worksheet.cell(row=row, column=name_col).value
                self.models_names[model_Index] = model_name

                for cmp_key in model_comparison_keys:
                    cmp_key_idx = list_of_keys.index(cmp_key)

                    tmp = models_worksheet.cell(row=row, column=cmp_key_idx+1).value
                    if (tmp == self.model[cmp_key]) or ((tmp is None) and (self.model[cmp_key] is None)):
                        ok_counter += 1

                if ok_counter == len(model_comparison_keys):
                    # fill model["UUID"] and ...
                    self.model["Index"] = models_worksheet.cell(row=row, column=1).value
                    self.model["UUID"] = models_worksheet.cell(row=row, column=UUID_col).value
                    self.model["name"] = models_worksheet.cell(row=row, column=name_col).value

        else:
            # Initialize xlsx file
            # Create a workbook and add a worksheet.
            Path(models_database_filename).parent.mkdir(parents=True, exist_ok=True)
            pd_workbook = xlsxwriter.Workbook(models_database_filename)
            pd_worksheet = pd_workbook.add_worksheet("Models")        

            # Add a bold format to use to highlight cells.
            bold = pd_workbook.add_format({'bold': True}) 
            #width = len(max(model.keys(), key=len)) + 1
            for i, key_name in enumerate(self.model.keys()):
                # format font (bold + size?)
                pd_worksheet.write(0, i, key_name, bold)
                # change column width
                width = len(key_name) + 1
                pd_worksheet.set_column(i, i, width)
            pd_workbook.close()
            no_of_models = 0

        if len(self.model["UUID"]) == 0:
            # TODO add model data to file
            # add model to models_data
            model_ID = uuid.uuid4()
            if len(self.model["name"]) == 0:
                # automatically generate name for model
                # model type + parameters length + no_of_classes +  classes_mode + learning rate
                model_type = self.model["model_type"]
                model_params_len = len(model_params["model_config"]["parameters"])
                if model_params_len > 1:
                    model_type = model_params["model_config"]["parameters"][0][0]
                    model_type += "-"
                    model_type += model_params["model_config"]["parameters"][1][0]
                no_of_classes =self.model["no_of_classes"]
                classes_mode = self.model["classes_mode"]
                learning_rate = self.model["learning_rate"]
                train_SNR_dB = self.model["training_SNR_dB"]
                reload = self.model["training_reload"]
                FB_to_CMPO_smoothing_factor = self.model["FB_to_CMPO_smoothing_factor"]
                if validation_SNR_str is None:
                    self.model["name"] = f"{model_type}({model_params_len})_{classes_mode}({no_of_classes})_lr_{learning_rate}_rl({reload})_trSNR_{train_SNR_dB}_smf_{FB_to_CMPO_smoothing_factor}"
                else:
                    self.model["name"] = f"{model_type}({model_params_len})_{classes_mode}({no_of_classes})_lr_{learning_rate}_rl({reload})_trSNR_{train_SNR_dB}_smf_{FB_to_CMPO_smoothing_factor}_valSNR_{validation_SNR_str}"

            self.model["UUID"] = model_ID

            no_of_models = no_of_models + 1
            self.model["Index"] = no_of_models

            self.models_names[no_of_models] = self.model["name"]

            # Update Excel file
            # Create a workbook and add a worksheet.
            # workbook = xlsxwriter.Workbook(models_database_filename)
            if workbook is None:
                workbook = openpyxl.load_workbook(models_database_filename)
            models_worksheet = workbook["Models"]

            # Add a bold format to use to highlight cells.
            #width = len(max(model.keys(), key=len)) + 1
            for i, key_name in enumerate(self.model.keys()):
                if type(self.model[key_name]) == int or type(self.model[key_name]) == float or type(self.model[key_name]) == bool or self.model[key_name] is None:
                    models_worksheet.cell(no_of_models+1, i+1, value=self.model[key_name]) 
                else:
                    models_worksheet.cell(no_of_models+1, i+1, value=str(self.model[key_name])) 
                # TODO format font?
                # change column width
                width = len(str(models_worksheet.cell(row=no_of_models+1, column=i+1).value)) + 1
                models_worksheet.column_dimensions[get_column_letter(i+1)].width = max([width, models_worksheet.column_dimensions[get_column_letter(i+1)].width])

            # add sheet for results of current model
            model_index = self.model["Index"]
            # model_sheet_name = f"{model_index}-{model_str}" exceeds 32 character Excel limit
            model_sheet_name = get_model_sheet_name(model_index, validation_signal_str)

            # =================================================== #
            model_worksheet = workbook.create_sheet(model_sheet_name)
            # Initialize sheet of model measured parameters
            model_worksheet.cell(self.UUID_ROW, 1, value="UUID") 
            model_worksheet.cell(self.UUID_ROW, 2, value=str(self.model["UUID"])) 
            model_worksheet.cell(self.NAME_ROW, 1, value="name") 
            model_worksheet.cell(self.NAME_ROW, 2, value=self.model["name"]) 
            model_worksheet.cell(self.MODEL_PATH_ROW, 1, value="model_path") 
            model_worksheet.cell(self.MODEL_PATH_ROW, 2, value=self.model["model_path"]) 
            model_worksheet.cell(self.VALIDATION_MODE_ROW, 1, value="validation_mode") 
            model_worksheet.cell(self.VALIDATION_MODE_ROW, 2, value=self.model["validation_mode"]) 

            model_worksheet.cell(self.EPOCH_NO_ROW, 2, value="epoch_no") 
            model_worksheet.cell(self.TRAIN_ACCU_ROW, 2, value="train. accur.") 
            model_worksheet.cell(self.TRAIN_SNR_ROW, 2, value="train_SNR_dB") 
            model_worksheet.cell(self.INFER_ACCU_ROW, 2, value="infer. accur.") 
            model_worksheet.cell(self.OUTLIERS_ROW, 2, value="outliers %") 
            model_worksheet.cell(self.MEAN_ROW, 2, value="mean") 
            model_worksheet.cell(self.STD_ROW, 2, value="std") 
            model_worksheet.cell(self.RMSE_ROW, 2, value="RMSE") 
            model_worksheet.cell(self.R2_ROW, 2, value="R^2") 
            model_worksheet.cell(self.MEAN_RELATIVE_ROW, 2, value="mean rel.") 
            model_worksheet.cell(self.STD_RELATIVE_ROW, 2, value="std rel.") 
            model_worksheet.cell(self.RMSE_RELATIVE_ROW, 2, value="RMSE rel.") 
            model_worksheet.cell(self.R2_RELATIVE_ROW, 2, value="R^2 rel.") 

            model_worksheet.cell(self.LABELS_ROW, 1, value="labels [Hz]") 
            model_worksheet.cell(self.LABELS_ROW, 2, value="F0_relative_ref") 
            model_worksheet.cell(self.LABELS_ROW, 3, value=self.F0_relative_ref) 

            name_strs = [str(model_worksheet.cell(row=i,column=1).value) for i in range(1,self.LABELS_ROW+1)]
            width = len(max(name_strs, key=len)) + 1
            model_worksheet.column_dimensions[get_column_letter(1)].width = width

            name_strs = [str(model_worksheet.cell(row=i,column=2).value) for i in range(self.EPOCH_NO_ROW,self.STD_ROW+1)]
            width = len(max(name_strs, key=len)) + 1
            model_worksheet.column_dimensions[get_column_letter(2)].width = width

            # add sheet for figures of current model
            # model_sheet_name = f"{model_index}-{model_str}" exceeds 32 character Excel limit
            
            model_figs_sheet_name = get_model_sheet_name(model_index, validation_signal_str) + "_Figs"

            # =================================================== #
            worksheet_Figs = workbook.create_sheet(model_figs_sheet_name)
            # # Initialize sheet of model figures
            # # TODO ???????????????????
            # # start from "C" column ("B" has labels)
            # # TODO max_col selection ???
            # values = Reference(model_worksheet, min_col=2, min_row=self.TRAIN_ACCU_ROW, max_col=max_epoch, max_row=self.INFER_ACCU_ROW)
            # chart = LineChart()
            # chart.title = self.models_names[model_index]
            # chart.style = 12 # 13
            # chart.y_axis.title = 'Size'
            # chart.x_axis.title = 'Test Number'
            # chart.add_data(values, titles_from_data=True, from_rows=True)
            # worksheet_Figs.add_chart(chart, "A1")

            workbook.save(models_database_filename)


        # # if models_data is None:
        # #     # create ne file
        # if model_data_changed == True:
        #     writer = pd.ExcelWriter(models_database_filename, engine='xlsxwriter')   
        #     models_data.to_excel(writer, sheet_name='Models', index=False)

        #     # Auto-adjust columns' width
        #     for column in models_data:
        #         column_width = max(models_data[column].astype(str).map(len).max(), len(column)) + 1
        #         col_idx = models_data.columns.get_loc(column)
        #         writer.sheets['Models'].set_column(col_idx, col_idx, column_width)        

        #     writer.save()

        return self.model

    def add_model_chart(self, model_index, validation_signal_str, max_epoch = 100):
        # a sheet with charts with data from model of given index

        # =================================================== #
        model_sheet_name = get_model_sheet_name(model_index, validation_signal_str)
        model_worksheet = self.workbook4stage[model_sheet_name]

        model_figs_sheet_name = get_model_sheet_name(model_index, validation_signal_str) + "_Figs"
        reuse_charts = False
        if model_figs_sheet_name in self.workbook4stage.sheetnames:
            worksheet_Fig = self.workbook4stage[model_figs_sheet_name]

            worksheet_Fig._charts.clear()
            # if len(worksheet_Fig._charts) > 0:
            #     accu_chart = worksheet_Fig._charts[0]
            #     outliers_chart = worksheet_Fig._charts[1]
            #     train_SNR_chart = worksheet_Fig._charts[2]
            #     RMSE_chart = worksheet_Fig._charts[3]
            #     R2_chart   = worksheet_Fig._charts[4]
            #     reuse_charts = True

        else:
            raise Exception(f"worksheet_Fig: {model_figs_sheet_name} does not exist")

        if reuse_charts == False:
            # =================================================== #
            # overall chart for training and inference accuracy
            epoch_values = Reference(model_worksheet, min_col=3, min_row=self.EPOCH_NO_ROW, max_col=3+max_epoch, max_row=self.EPOCH_NO_ROW)

            accu_chart = ScatterChart()
            if validation_signal_str is None:
                accu_chart.title = self.models_names[model_index]
            else:
                accu_chart.title = self.models_names[model_index] + f" ({validation_signal_str})"
            accu_chart.style = 12 # 13
            accu_chart.y_axis.title = '%'
            accu_chart.x_axis.title = 'Epoch number'
            accu_chart.height = 10 # default is 7.5
            accu_chart.width = 30 # default is 15

            train_values = Reference(model_worksheet, min_col=3, min_row=self.TRAIN_ACCU_ROW, max_col=3+max_epoch, max_row=self.TRAIN_ACCU_ROW)
            train_series = Series(train_values, epoch_values, title="train. accu.") #  model_sheet_name)
            train_series.graphicalProperties.line.solidFill = self.get_color(1)
            train_series.graphicalProperties.line.width = 20000
            train_series.smooth = False
            train_series.display_blank = 'span'

            infer_values = Reference(model_worksheet, min_col=3, min_row=self.INFER_ACCU_ROW, max_col=3+max_epoch, max_row=self.INFER_ACCU_ROW)
            infer_series = Series(infer_values, epoch_values, title="infer. accu.")
            infer_series.graphicalProperties.line.solidFill = self.get_color(2)
            # lineProp = drawing.line.LineProperties(prstDash='dash')
            infer_series.graphicalProperties.line.dashStyle = 'dash'
            infer_series.graphicalProperties.line.width = 20000
            infer_series.smooth = False
            infer_series.display_blank = 'span'

            accu_chart.append(train_series)
            accu_chart.append(infer_series)

            # ====================================================== #
            outliers_chart = ScatterChart()
            if validation_signal_str is None:
                outliers_chart.title = self.models_names[model_index]
            else:
                outliers_chart.title = self.models_names[model_index] + f" ({validation_signal_str})"
            outliers_chart.style = 12 # 13
            outliers_chart.y_axis.title = '%'
            outliers_chart.x_axis.title = 'Epoch number'
            outliers_chart.height = 10 # default is 7.5
            outliers_chart.width = 30 # default is 15

            outliers_values = Reference(model_worksheet, min_col=3, min_row=self.OUTLIERS_ROW, max_col=3+max_epoch, max_row=self.OUTLIERS_ROW)
            outliers_series = Series(outliers_values, epoch_values, title="outliers") #  model_sheet_name)
            outliers_series.graphicalProperties.line.solidFill = self.get_color(1)
            outliers_series.graphicalProperties.line.width = 20000
            outliers_series.smooth = False
            outliers_series.display_blank = 'span'

            outliers_chart.append(outliers_series)

            # ====================================================== #
            train_SNR_chart = ScatterChart()
            if validation_signal_str is None:
                train_SNR_chart.title = self.models_names[model_index]
            else:
                train_SNR_chart.title = self.models_names[model_index] + f" ({validation_signal_str})"
            train_SNR_chart.style = 12 # 13
            train_SNR_chart.y_axis.title = 'dB'
            train_SNR_chart.x_axis.title = 'Epoch number'
            train_SNR_chart.height = 10 # default is 7.5
            train_SNR_chart.width = 30 # default is 15

            train_SNR_values = Reference(model_worksheet, min_col=3, min_row=self.TRAIN_SNR_ROW, max_col=3+max_epoch, max_row=self.TRAIN_SNR_ROW)
            train_SNR_series = Series(train_SNR_values, epoch_values, title="SNR") #  model_sheet_name)
            train_SNR_series.graphicalProperties.line.solidFill = self.get_color(1)
            train_SNR_series.graphicalProperties.line.width = 20000
            train_SNR_series.smooth = False
            train_SNR_series.display_blank = 'span'

            train_SNR_chart.append(train_SNR_series)

            # ====================================================== #
            RMSE_chart = ScatterChart()
            if validation_signal_str is None:
                RMSE_chart.title = self.models_names[model_index]
            else:
                RMSE_chart.title = self.models_names[model_index] + f" ({validation_signal_str})"
            RMSE_chart.style = 12 # 13
            RMSE_chart.y_axis.title = 'Hz'
            RMSE_chart.x_axis.title = 'Epoch number'
            RMSE_chart.height = 10 # default is 7.5
            RMSE_chart.width = 30 # default is 15

            RMSE_values = Reference(model_worksheet, min_col=3, min_row=self.RMSE_ROW, max_col=3+max_epoch, max_row=self.RMSE_ROW)
            RMSE_series = Series(RMSE_values, epoch_values, title="RMSE") #  model_sheet_name)
            RMSE_series.graphicalProperties.line.solidFill = self.get_color(1)
            RMSE_series.graphicalProperties.line.width = 20000
            RMSE_series.smooth = False
            RMSE_series.display_blank = 'span'

            mean_values = Reference(model_worksheet, min_col=3, min_row=self.MEAN_ROW, max_col=3+max_epoch, max_row=self.MEAN_ROW)
            mean_series = Series(mean_values, epoch_values, title="mean") #  model_sheet_name)
            mean_series.graphicalProperties.line.solidFill = self.get_color(3)
            mean_series.graphicalProperties.line.width = 20000
            mean_series.smooth = False
            mean_series.display_blank = 'span'

            std_values = Reference(model_worksheet, min_col=3, min_row=self.STD_ROW, max_col=3+max_epoch, max_row=self.STD_ROW)
            std_series = Series(std_values, epoch_values, title="std") #  model_sheet_name)
            std_series.graphicalProperties.line.solidFill = self.get_color(2)
            std_series.graphicalProperties.line.dashStyle = 'dash'
            std_series.graphicalProperties.line.width = 20000
            std_series.smooth = False
            std_series.display_blank = 'span'

            RMSE_chart.append(RMSE_series)
            RMSE_chart.append(mean_series)
            RMSE_chart.append(std_series)

            # ====================================================== #
            R2_chart = ScatterChart()
            if validation_signal_str is None:
                R2_chart.title = self.models_names[model_index]
            else:
                R2_chart.title = self.models_names[model_index] + f" ({validation_signal_str})"
            R2_chart.style = 12 # 13
            R2_chart.y_axis.title = ''
            R2_chart.x_axis.title = 'Epoch number'
            R2_chart.height = 10 # default is 7.5
            R2_chart.width = 30 # default is 15

            R2_values = Reference(model_worksheet, min_col=3, min_row=self.R2_ROW, max_col=3+max_epoch, max_row=self.R2_ROW)
            R2_series = Series(R2_values, epoch_values, title="R^2") #  model_sheet_name)
            R2_series.graphicalProperties.line.solidFill = self.get_color(1)
            R2_series.graphicalProperties.line.width = 20000
            R2_series.smooth = False
            R2_series.display_blank = 'span'

            R2_chart.append(R2_series)


        accu_chart.display_blanks = 'span'
        outliers_chart.display_blanks = 'span'
        train_SNR_chart.display_blanks = 'span'
        RMSE_chart.display_blanks = 'span'
        R2_chart.display_blanks = 'span'
            
        if reuse_charts == False:
            accu_chart.y_axis.scaling.min = 0
            # RMSE_chart.y_axis.scaling.min = 0
            # R2_chart.y_axis.scaling.min = 0

            worksheet_Fig.add_chart(accu_chart, "A1")
            worksheet_Fig.add_chart(outliers_chart, "A21")
            worksheet_Fig.add_chart(train_SNR_chart, "A41")
            worksheet_Fig.add_chart(RMSE_chart, "A61")
            worksheet_Fig.add_chart(R2_chart, "A81")

    def add_models_chart(self, validation_signal_str, max_epoch = 100):
        # a single sheet with charts with data from all models
        # separate for training data and for inference data  

        # =================================================== #
        models_fig_sheet_name = "Models_Figs"
        reuse_charts = False
        if models_fig_sheet_name in self.workbook4stage.sheetnames:
            worksheet_Figs = self.workbook4stage[models_fig_sheet_name]

            worksheet_Figs._charts.clear()
            # train_chart = worksheet_Figs._charts[0]
            # infer_chart = worksheet_Figs._charts[1]
            # outliers_chart = worksheet_Figs._charts[2]
            # RMSE_chart  = worksheet_Figs._charts[3]
            # R2_chart    = worksheet_Figs._charts[4]
            # reuse_charts = True

        else:
            worksheet_Figs = self.workbook4stage.create_sheet(models_fig_sheet_name, 1)

        if reuse_charts == False:
            # =================================================== #
            # chart for training accuracy
            train_chart = ScatterChart()
            train_chart.title = "Models train. accur." # self.model["name"]
            train_chart.style = 12 # 13
            train_chart.y_axis.title = '%'
            train_chart.x_axis.title = 'Epoch number'
            train_chart.height = 10 # default is 7.5
            train_chart.width = 50 # default is 15
            # =================================================== #
            # chart for inference accuracy
            infer_chart = ScatterChart()
            infer_chart.title = "Models infer. accur." # self.model["name"]
            infer_chart.style = 12 # 13
            infer_chart.y_axis.title = '%'
            infer_chart.x_axis.title = 'Epoch number'
            infer_chart.height = 10 # default is 7.5
            infer_chart.width = 50 # default is 15
            # =================================================== #
            # chart for outliers
            outliers_chart = ScatterChart()
            outliers_chart.title = "Models outliers" # self.model["name"]
            outliers_chart.style = 12 # 13
            outliers_chart.y_axis.title = '%'
            outliers_chart.x_axis.title = 'Epoch number'
            outliers_chart.height = 10 # default is 7.5
            outliers_chart.width = 50 # default is 15
            # =================================================== #
            # chart for RMSE
            RMSE_chart = ScatterChart()
            RMSE_chart.title = "Models RMSE" 
            RMSE_chart.style = 12 # 13
            RMSE_chart.y_axis.title = 'Hz'
            RMSE_chart.x_axis.title = 'Epoch number'
            RMSE_chart.height = 10 # default is 7.5
            RMSE_chart.width = 50 # default is 15
            # =================================================== #
            # chart for R^2
            R2_chart = ScatterChart()
            R2_chart.title = "Models R^2" 
            R2_chart.style = 12 # 13
            R2_chart.y_axis.title = '%'
            R2_chart.x_axis.title = 'Epoch number'
            R2_chart.height = 10 # default is 7.5
            R2_chart.width = 50 # default is 15

        # get model indexes fromm all sheets in xlsx and add Series from them to charts
        for model_sheet_name in self.workbook4stage.sheetnames:
            #  get model_index and validation_signal_str
            model_index, validation_signal_str = get_params_from_model_sheet_name(model_sheet_name)
            if model_index is not None:

        # model_index = 1
        # model_sheet_name = get_model_sheet_name(model_index, validation_signal_str)
        # while model_sheet_name in self.workbook4stage.sheetnames:
                model_worksheet = self.workbook4stage[model_sheet_name]

                # =================================================== #
                if validation_signal_str is None:
                    series_title = self.models_names[model_index]
                else:
                    series_title = self.models_names[model_index] + f" ({validation_signal_str})"
                # ? sprawdzić po train_chart.series.title ??
                train_series_idx = None
                for series_idx in range(len(train_chart.series)):
                    series = train_chart.series[series_idx]
                    if series.title.value == series_title:
                        train_series_idx = series_idx
                infer_series_idx = None
                for series_idx in range(len(infer_chart.series)):
                    series = infer_chart.series[series_idx]
                    if series.title.value == series_title:
                        infer_series_idx = series_idx
                outliers_series_idx = None
                for series_idx in range(len(outliers_chart.series)):
                    series = outliers_chart.series[series_idx]
                    if series.title.value == series_title:
                        outliers_series_idx = series_idx
                RMSE_series_idx = None
                for series_idx in range(len(RMSE_chart.series)):
                    series = RMSE_chart.series[series_idx]
                    if series.title.value == series_title:
                        RMSE_series_idx = series_idx
                R2_series_idx = None
                for series_idx in range(len(R2_chart.series)):
                    series = R2_chart.series[series_idx]
                    if series.title.value == series_title:
                        R2_series_idx = series_idx

                epoch_values = Reference(model_worksheet, min_col=3, min_row=self.EPOCH_NO_ROW, max_col=3+max_epoch, max_row=self.EPOCH_NO_ROW)
                if train_series_idx == None:
                    train_values = Reference(model_worksheet, min_col=3, min_row=self.TRAIN_ACCU_ROW, max_col=3+max_epoch, max_row=self.TRAIN_ACCU_ROW)
                    # train_chart.add_data(train_values, titles_from_data=True, from_rows=True)
                    #
                    # train_chart.add_data(train_values, titles_from_data=False, from_rows=True)
                    train_series = Series(train_values, epoch_values, title=series_title) #  model_sheet_name)
                    train_series.graphicalProperties.line.solidFill = self.get_color(model_index)
                    train_series.graphicalProperties.line.width = 20000
                    train_series.smooth = False
                    train_series.display_blank = 'span'
                    train_chart.append(train_series)
                
                if infer_series_idx == None:
                    infer_values = Reference(model_worksheet, min_col=3, min_row=self.INFER_ACCU_ROW, max_col=3+max_epoch, max_row=self.INFER_ACCU_ROW)
                    # infer_chart.add_data(infer_values, titles_from_data=True, from_rows=True)
                    infer_series = Series(infer_values, epoch_values, title=series_title)
                    infer_series.graphicalProperties.line.solidFill = self.get_color(model_index)
                    infer_series.graphicalProperties.line.width = 20000
                    infer_series.smooth = False
                    infer_series.display_blank = 'span'
                    infer_chart.append(infer_series)
                
                if outliers_series_idx == None:
                    outliers_values = Reference(model_worksheet, min_col=3, min_row=self.OUTLIERS_ROW, max_col=3+max_epoch, max_row=self.OUTLIERS_ROW)
                    # infer_chart.add_data(infer_values, titles_from_data=True, from_rows=True)
                    outliers_series = Series(outliers_values, epoch_values, title=series_title)
                    outliers_series.graphicalProperties.line.solidFill = self.get_color(model_index)
                    outliers_series.graphicalProperties.line.width = 20000
                    outliers_series.smooth = False
                    outliers_series.display_blank = 'span'
                    outliers_chart.append(outliers_series)
                
                if RMSE_series_idx == None:
                    RMSE_values = Reference(model_worksheet, min_col=3, min_row=self.RMSE_ROW, max_col=3+max_epoch, max_row=self.RMSE_ROW)
                    # infer_chart.add_data(infer_values, titles_from_data=True, from_rows=True)
                    RMSE_series = Series(RMSE_values, epoch_values, title=series_title)
                    RMSE_series.graphicalProperties.line.solidFill = self.get_color(model_index)
                    RMSE_series.graphicalProperties.line.width = 20000
                    RMSE_series.smooth = False
                    RMSE_series.display_blank = 'span'
                    RMSE_chart.append(RMSE_series)
                
                if R2_series_idx == None:
                    R2_values = Reference(model_worksheet, min_col=3, min_row=self.R2_ROW, max_col=3+max_epoch, max_row=self.R2_ROW)
                    # infer_chart.add_data(infer_values, titles_from_data=True, from_rows=True)
                    R2_series = Series(R2_values, epoch_values, title=series_title)
                    R2_series.graphicalProperties.line.solidFill = self.get_color(model_index)
                    R2_series.graphicalProperties.line.width = 20000
                    R2_series.smooth = False
                    R2_series.display_blank = 'span'
                    R2_chart.append(R2_series)

                self.add_model_chart(model_index, validation_signal_str, max_epoch)

            # model_index += 1
            # model_sheet_name = get_model_sheet_name(model_index, validation_signal_str)

        train_chart.display_blanks = 'span'
        infer_chart.display_blanks = 'span'
        outliers_chart.display_blanks = 'span'
        RMSE_chart.display_blanks = 'span'
        R2_chart.display_blanks = 'span'

        if reuse_charts == False:
            train_chart.y_axis.scaling.min = 0
            infer_chart.y_axis.scaling.min = 0
            outliers_chart.y_axis.scaling.min = 0
            # RMSE_chart.y_axis.scaling.min = 0
            # R2_chart.y_axis.scaling.min = 0

            worksheet_Figs.add_chart(train_chart, "A1")
            worksheet_Figs.add_chart(infer_chart, "A21")
            worksheet_Figs.add_chart(outliers_chart, "A41")
            worksheet_Figs.add_chart(RMSE_chart, "A61")
            worksheet_Figs.add_chart(R2_chart, "A81")



# def load_inference_data(validation_dataset_subpath, training_setup_subpath, epoch_str):
#     # TODO load inference data for given validation_dataset_subpath and training_setup_subpath
#     data_path = Path("/output_data/inference", validation_dataset_subpath, training_setup_subpath)
#     # find all 

def load_inference_data(epochs_infer_save_path, epoch_strings, infer_data_set_idx, model_params):
    # load infered estimates and Fo_ref_labels
    F0_ref_filename = epochs_infer_save_path.parent / "infer_F0_ref"
    if Path(F0_ref_filename).with_suffix(".npz").is_file():
        loaded_F0_ref = np.load(Path(F0_ref_filename).with_suffix(".npz"))

    F0_ref_indexes_filename = epochs_infer_save_path / "infer_F0_ref_indexes"
    if Path(F0_ref_indexes_filename).with_suffix(".npz").is_file():
        loaded = np.load(Path(F0_ref_indexes_filename).with_suffix(".npz"))
    F0_ref_indexes = loaded["F0_ref_indexes"]

    infer_results = {}
    infer_results["F0_ref"] = loaded_F0_ref["F0_ref"][F0_ref_indexes]

    infer_filename = epochs_infer_save_path / ("infer_data_" + epoch_strings[infer_data_set_idx])
    if Path(infer_filename).with_suffix(".npz").is_file():
        loaded = np.load(Path(infer_filename).with_suffix(".npz"), allow_pickle=True)

        if "training_SNR_dB" in loaded:
            training_SNR_dB = loaded["training_SNR_dB"]
            # print(type(training_SNR_dB))
            # print(isinstance(training_SNR_dB, np.ndarray))
            # print(training_SNR_dB)
            # print(training_SNR_dB.item())
            if isinstance(training_SNR_dB, np.ndarray):
                training_SNR_dB = training_SNR_dB.item()
            # if isinstance(training_SNR_dB, float):
            #     training_SNR_dB = None
        else:
            training_SNR_dB = None

        if training_SNR_dB is None:
            # try to generate based on entropy if cfg file is available; will fail if the seed file have changed
            current_epoch_idx = int(epoch_strings[infer_data_set_idx][6:])

            if "training_dataset_options" in model_params:
                if model_params["training_dataset_options"]["rng_state_filename"] is not None:
                    rng_state_filename = Path(model_params["model_path"]).parent / model_params["training_dataset_options"]["rng_state_filename"]
                    reload_step = model_params["training_dataset_options"]["reload"]
                    SNR_dB_range = model_params["training_dataset_options"]["SNR_dB"]

                    if reload_step == 0:
                        rng_epoch_idx = 0
                    elif reload_step == 1:
                        rng_epoch_idx = current_epoch_idx
                    else:
                        rng_epoch_idx = current_epoch_idx - (current_epoch_idx % reload_step)
                    SNR_index = current_epoch_idx

                    vs  = vowels_synthesizer(rng_epoch_idx, rng_state_filename)
                    training_SNR_dB = get_SNR_dB(rng_state_filename, SNR_index, SNR_dB_range, vs)

        # print(training_SNR_dB)

        infer_results["F0_est"] = loaded["F0_est"]
        infer_results["F0_ref_labels"] = loaded["F0_ref_labels"]
        # loaded["model_training_accuracy"][()] <= alternative to loaded["model_training_accuracy"].item()
        model_train_accuracy = loaded["model_training_accuracy"].item()["train_accuracy"]
        model_accuracy_table = loaded["model_training_accuracy"].item()["accuracy_table"]
        model_batch_size = loaded["model_training_accuracy"].item()["batch_size"]

    return infer_results, model_train_accuracy, model_accuracy_table, model_batch_size, training_SNR_dB

def process_folder(main_trained_models_folder, pf_cfg, validation_SNR_str, datetime_threshold_str = None):

    # TODO read configuration from json file
    # # configuration for the results analysis
    # results_analysis_config = inferer_options["results_analysis_config"]
    # do_save = results_analysis_config["do_save"] # False to skip saving output files
    # do_draw_epochs = results_analysis_config["do_draw_epochs"]
    # if do_draw_epochs:
    #     draw_epoch_step = results_analysis_config["draw_epoch_step"]
    # else:
    #     draw_epoch_step = 100
    # pause_after_model = results_analysis_config["pause_after_model"]
    # # # TODO add starting epoch or vector with epoch indexes

    do_split_xlsx_by_classes_mode = pf_cfg["do_split_xlsx_by_classes_mode"]
    xls_store_per_class_data = pf_cfg["xls_store_per_class_data"]
    do_pause = pf_cfg["do_pause"]
    do_split_by_validation_signals = pf_cfg["do_split_by_validation_signals"]
    if do_split_by_validation_signals == True:
        validation_signal_str = None
    else:
        validation_signal_str = pf_cfg["validation_signal_folder"]
    validation_signal_folder = pf_cfg["validation_signal_folder"]

    use_date_time_str = pf_cfg["use_date_time_str"]
    do_split_by_model= pf_cfg["do_split_by_model"]
    main_folder = pf_cfg["main_folder"]
    validation_signals_tag_str = pf_cfg["validation_signals_tag_str"]
    datetime_str = pf_cfg["datetime_str"]

    # restart_needed = False

    # training_setup_subpaths = [ "MLP2_lin_lin_lin_0_lin_N=100_False_synth_reload_1_SNR_40" ]
    if not Path("./output_data/inference", main_trained_models_folder).is_dir():
        print(f"main_trained_models_folder: {main_trained_models_folder} does not exist")
        return False

    training_setup_subpaths = [ f.path for f in os.scandir(Path("./output_data/inference", main_trained_models_folder)) if (f.is_dir()) ]

    md2excel = ModelData2Excel()
    for setup_idx, full_training_setup_subpath in enumerate(training_setup_subpaths):
        training_setup_subpath = PurePath(full_training_setup_subpath).name
        print(f"Processing {training_setup_subpath} ...")

        if use_date_time_str == True:
            if do_split_by_validation_signals == True:
                if do_split_by_model == True:
                    models_filename_tag_str = f"{main_folder}/{training_setup_subpath}/{validation_signal_folder}/{validation_signals_tag_str}_{datetime_str}_"
                else:
                    models_filename_tag_str = f"{main_folder}/{validation_signal_folder}/{validation_signals_tag_str}_{datetime_str}_"
            else:
                if do_split_by_model == True:
                    models_filename_tag_str = f"{main_folder}/{training_setup_subpath}/{validation_signals_tag_str}_{datetime_str}_"
                else:
                    models_filename_tag_str = f"{main_folder}/{validation_signals_tag_str}_{datetime_str}_"
        else:
            if do_split_by_validation_signals == True:
                if do_split_by_model == True:
                    models_filename_tag_str = f"{main_folder}/{training_setup_subpath}/{validation_signal_folder}/{validation_signals_tag_str}_"
                else:
                    models_filename_tag_str = f"{main_folder}/{validation_signal_folder}/{validation_signals_tag_str}_"
            else:
                if do_split_by_model == True:
                    models_filename_tag_str = f"{main_folder}/{training_setup_subpath}/{validation_signals_tag_str}_"
                else:
                    models_filename_tag_str = f"{main_folder}/{validation_signals_tag_str}_"

        restart_needed = False 

        # =========================================================================================================
        epoch_strings, epoch_indexes, epochs_infer_save_path = find_inference_data(validation_dataset_subpath=main_trained_models_folder, 
                                                                                training_setup_subpath=training_setup_subpath)
        # TODO load network model parameters
        model_params_filename = epochs_infer_save_path / "model_params.json"
        if Path(model_params_filename).is_file():
            # load model parameter in json file
            with open(model_params_filename,'r+') as f:
                model_params = json.load(f)            
                # model_params = {
                #         "model_path": model_path,
                #         "model_config": model_config,
                #         "labels_transformer_options": labels_transformer_options,
                #         "checkpoint_trainer_config": checkpoint_trainer_config
                #     }
                if "AFE" not in model_params["dataset_options"]:
                    AFE_cfg = {"filter_bank_mode": "ConstQ", 
                                "filter_bank_smoothing_mode": "delay",
                                "CMPO_smoothing_mode": "none",
                                "FB_to_CMPO_smoothing_factor": 1.0}    
                    logging.info(r"Warning: dataset_options[\"AFE\"] does not exist, using default AFE config")
                    model_params["dataset_options"]["AFE"] = AFE_cfg
                else:
                    if "FB_to_CMPO_smoothing_factor" not in model_params["dataset_options"]["AFE"]:
                        model_params["dataset_options"]["AFE"]["FB_to_CMPO_smoothing_factor"] = -1.0 # signal that FB_to_CMPO_smoothing_factor wasn't defined
                        logging.info(r"Warning: dataset_options[\"AFE\"] does not contain \"FB_to_CMPO_smoothing_factor\" setting it to -1.0")

        else:
            # TODO load data from epoch data and save to model_params_filename
            raise Exception(f"missing {model_params_filename} file")

        no_of_classes = model_params["model_config"]["no_of_classes"]
        no_of_inputs = model_params["model_config"]["no_of_inputs"]
        if do_split_xlsx_by_classes_mode == True:
            classes_mode = model_params["labels_transformer_options"]["mode"]
            models_database_filename = Path("./output_data/inference/xlsx", f"{models_filename_tag_str}models_database_in_{no_of_inputs}_classes_{no_of_classes}_{classes_mode}.xlsx")
        else:
            models_database_filename = Path("./output_data/inference/xlsx", f"{models_filename_tag_str}models_database_in_{no_of_inputs}_classes_{no_of_classes}.xlsx")

        if "training_data_options" not in model_params:
            # quick fix
            reload_str_pos = model_params_filename.parent.name.find("reload_")
            if reload_str_pos == -1:
                reload_str = "1"
            else:
                reload_str = model_params_filename.parent.name[reload_str_pos+7:]
                reload_str_pos = reload_str.find("_")
                reload_str = reload_str[:reload_str_pos]

            # SNR_str = model_params_filename.parent.name[-2:]
            SNR_str_pos = model_params_filename.parent.name.find("SNR_")
            if SNR_str_pos == -1:
                SNR_str_1 = "40"
                SNR_str_2 = "40"
            else:
                SNR_str = model_params_filename.parent.name[SNR_str_pos+4:]
                SNR_str_pos = SNR_str.find("_")
                if SNR_str_pos == -1:
                    SNR_str = SNR_str[:]
                else:
                    SNR_str = SNR_str[:SNR_str_pos]
                
                SNR_str_pos = SNR_str[1:].find("-") 
                if SNR_str_pos == -1:
                    SNR_str_1 = SNR_str
                    SNR_str_2 = SNR_str
                else:
                    SNR_str_1 = SNR_str[:SNR_str_pos+1]
                    SNR_str_2 = SNR_str[SNR_str_pos+2:]

                
            model_params["training_data_options"] = {"mode": "synthesized",
                                                    "SNR_dB": [int(SNR_str_1), int(SNR_str_2)],
                                                    "reload": int(reload_str),
                                                    }

        # check if models is in database if no, add it
        md2excel.find_model_in_database(model_params, models_database_filename, validation_signal_str, validation_SNR_str)
        update_excel_on_load = True # TODO add to config parameter

        analysis_data_filename = epochs_infer_save_path / epochs_infer_save_path.name


        if datetime_threshold_str is not None:
            if analysis_data_filename.with_suffix(analysis_data_filename.suffix + ".npz").is_file() == True:
                file_mtime_secs = Path(analysis_data_filename.with_suffix(analysis_data_filename.suffix + ".npz")).stat().st_mtime
                # tmp3 = time.gmtime(file_mtime) # [secs]
                
                # # os.path.getmtime()
                # # tmp1 = time.time()
                # datetime_threshold = time.strptime("2022-03-09 10:00", "%Y-%m-%d %H:%M")
                datetime_threshold = time.strptime(datetime_threshold_str, "%Y-%m-%d %H:%M")
                datetime_threshold_secs = time.mktime(datetime_threshold)
                
                file_delta_time = file_mtime_secs - datetime_threshold_secs

                if file_delta_time < 0:
                    # previous file is too old => delete it
                    filename_to_delete = analysis_data_filename.with_suffix(analysis_data_filename.suffix + ".npz")
                    # os.remove(filename_to_delete)
                    filename_to_delete.unlink()

        # if restart_needed == True:
        #     # delete output file it it exists
        #     if analysis_data_filename.with_suffix(analysis_data_filename.suffix + ".npz").is_file() == True:
        #         filename_to_delete = analysis_data_filename.with_suffix(analysis_data_filename.suffix + ".npz")
        #         # os.remove(filename_to_delete)
        #         filename_to_delete.unlink()
        #     restart_needed = False 

        if analysis_data_filename.with_suffix(analysis_data_filename.suffix + ".npz").is_file() == False:
            epoch_dp = EpochDataProcessor() # initialization

            print(f"Processing: {epochs_infer_save_path}")

            classes_labels = np.array(IFEDataset.load_classes_labels(epochs_infer_save_path))
            if model_params["labels_transformer_options"]["use_F0_too_low_class"] == False and classes_labels[0] == 0:
                classes_labels = classes_labels[1:]
            md2excel.add_model_classes_labels(classes_labels, validation_signal_str)

            if epoch_dp.get_classes_labels() is None:
                epoch_dp.set_classes_labels(classes_labels)


            draw_epoch_step = 6
            do_draw_epochs = True

            # model_accuracy_table_all = np.zeros(shape=(0,))
            # reference_model_accuracy_table_size = None

            start_time = timer()

            infer_accuracy = np.zeros_like(epoch_indexes, dtype="float")
            with tqdm(total=len(epoch_indexes), dynamic_ncols=True) as progressbar:
                for infer_data_set_idx, model_epoch_idx in enumerate(epoch_indexes):
                    # load infered estimates and Fo_ref_labels
                    infer_results, model_train_accuracy, model_accuracy_table, model_batch_size, trained_SNR_dB = load_inference_data(epochs_infer_save_path, epoch_strings, infer_data_set_idx, model_params)

                    # recalculate accuracy_table
                    epoch_dp.process_accuracy(model_epoch_idx, model_accuracy_table, accu_skip_step = 50, model_batch_size = model_batch_size)

                    # # if (model_epoch % draw_epoch_step == 0) or (model_epoch == self.no_of_epochs-1):
                    # if (model_epoch % draw_epoch_step == 0) or (infer_data_set_idx == len(epoch_indexes)-1):
                    #     epoch_dp.append_epoch_data(infer_results, model_train_accuracy, model_epoch, do_draw_epochs)
                    # else:
                    #     epoch_dp.append_epoch_data(infer_results, model_train_accuracy, model_epoch, False)
                    epoch_dp.append_epoch_data(infer_results, model_train_accuracy, trained_SNR_dB, model_epoch_idx, False)

                    # save to excel
                    #md2excel.stage_model_parameters(model_epoch_idx, training_accuracy=model_train_accuracy, inference_accuracy=epoch_dp.exact_percentage[-1])
                    # alternatively use epoch_dp.epoch_indexes to find infer_data_set_idx
                    md2excel.stage_model_parameters(model_epoch_idx, 
                        validation_signal_str=validation_signal_str,
                        training_accuracy=epoch_dp.train_accuracy_all[infer_data_set_idx],
                        inference_accuracy=epoch_dp.exact_percentage[infer_data_set_idx],
                        outliers_percentage=epoch_dp.outliers_percentage[infer_data_set_idx], 
                        error_without_outliers_mean=epoch_dp.F0_error_mean_all[infer_data_set_idx], 
                        error_without_outliers_std=epoch_dp.F0_error_std_all[infer_data_set_idx],
                        error_without_outliers_mean_relative = epoch_dp.F0_error_relative_mean_all[infer_data_set_idx], 
                        error_without_outliers_std_relative = epoch_dp.F0_error_relative_std_all[infer_data_set_idx],
                        trained_SNR_dB=epoch_dp.trained_SNR_dB_all[infer_data_set_idx])
                    progressbar.update(1)

            if xls_store_per_class_data == True:
                md2excel.add_parameters_per_class(validation_signal_str,
                                                  epoch_dp.no_of_data_per_class_all, epoch_dp.no_of_outliers_per_class_all, 
                                                  epoch_dp.est_mean_error_per_class_all, epoch_dp.est_std_error_per_class_all)

            # max_epoch = len(epoch_dp.train_accuracy_all)
            md2excel.commit_model_parameters(validation_signal_str)
            end_time = timer()
            print(f"elapsed time:{(end_time - start_time):.2f}") # Time in seconds, e.g. 5.38091952400282

                
            processed_accuracy = epoch_dp.get_processed_accuracy()

            # save data
            # np.savez_compressed(analysis_data_filename, processed_accuracy=processed_accuracy, epoch_dp=epoch_dp) 
            np.savez_compressed(analysis_data_filename, epoch_dp=epoch_dp) 

            print(f"Saved: {analysis_data_filename}")

        else:
            # loaded = np.load(analysis_data_filename.with_suffix(analysis_data_filename.suffix + ".npz"), allow_pickle=True)
            with open(analysis_data_filename.with_suffix(analysis_data_filename.suffix + ".npz"), 'rb') as f: # this variant closes file
                loaded = np.load(f, allow_pickle=True)
                epoch_dp = loaded["epoch_dp"][()]
                if not hasattr(epoch_dp, 'trained_SNR_dB_all'):
                    raise Exception("epoch_dp data have no trained_SNR_dB_all entry")

            print(f"Loaded: {epochs_infer_save_path}")
            if len(epoch_dp.train_accuracy_all) < len(epoch_strings):
                filename_to_delete = analysis_data_filename.with_suffix(analysis_data_filename.suffix + ".npz")
                # os.remove(filename_to_delete)
                filename_to_delete.unlink()
                restart_needed = True
                print(f"Not all data available: processing skipped + deleted {filename_to_delete}")

                return restart_needed

            if do_pause == True:
                # load last epoch data (just for drawing)
                infer_results, model_train_accuracy, model_accuracy_table, model_batch_size, trained_SNR_dB = load_inference_data(epochs_infer_save_path, epoch_strings, -1, model_params)


            if update_excel_on_load:
                # save to excel
                classes_labels = epoch_dp.get_classes_labels()
                if epoch_dp.get_classes_labels() is not None:
                    md2excel.add_model_classes_labels(classes_labels, validation_signal_str)

                
                if not hasattr(epoch_dp, 'no_of_data_per_class_all'):
                    filename_to_delete = analysis_data_filename.with_suffix(analysis_data_filename.suffix + ".npz")
                    # os.remove(filename_to_delete)
                    filename_to_delete.unlink()
                    restart_needed = True
                    print(f"Incorrect format: processing skipped + deleted {filename_to_delete}")
                else:
                    # save to excel
                    with tqdm(total=len(epoch_dp.train_accuracy_all), dynamic_ncols=True) as progressbar:
                        # for model_epoch_idx in range(len(epoch_dp.train_accuracy_all)):
                        for infer_data_set_idx, model_epoch_idx in enumerate(epoch_dp.epoch_indexes):
                            md2excel.stage_model_parameters(model_epoch_idx, 
                                validation_signal_str = validation_signal_str,
                                training_accuracy=epoch_dp.train_accuracy_all[infer_data_set_idx],
                                inference_accuracy=epoch_dp.exact_percentage[infer_data_set_idx],
                                outliers_percentage=epoch_dp.outliers_percentage[infer_data_set_idx], 
                                error_without_outliers_mean=epoch_dp.F0_error_mean_all[infer_data_set_idx], 
                                error_without_outliers_std=epoch_dp.F0_error_std_all[infer_data_set_idx],
                                error_without_outliers_mean_relative=epoch_dp.F0_error_relative_mean_all[infer_data_set_idx],
                                error_without_outliers_std_relative=epoch_dp.F0_error_relative_std_all[infer_data_set_idx],
                                trained_SNR_dB=epoch_dp.trained_SNR_dB_all[infer_data_set_idx])
                            progressbar.update(1)

                    if xls_store_per_class_data == True:
                        md2excel.add_parameters_per_class(validation_signal_str,
                                                        epoch_dp.no_of_data_per_class_all, epoch_dp.no_of_outliers_per_class_all, 
                                                        epoch_dp.est_mean_error_per_class_all, epoch_dp.est_std_error_per_class_all)

                    # max_epoch = len(epoch_dp.train_accuracy_all)
                    md2excel.commit_model_parameters(validation_signal_str)


        print(f"Finished processing {training_setup_subpath}")

        if do_pause == True:
            figs = draw_analyzis_results(epoch_dp, infer_results, setup_idx)

            # for i in plt.get_fignums():
            #     plt.figure(i).text(0, 1, training_setup_subpath, fontsize=8, verticalalignment='top')
            #     plt.figure(i).show()
            #     plt.figure(i).canvas.flush_events()
            for i in figs.keys():
                tmp = figs[i].canvas.manager.get_window_title()
                figs[i].canvas.manager.set_window_title(f"{setup_idx}: {tmp}")
                figs[i].text(0, 1, training_setup_subpath, fontsize=8, verticalalignment='top')
                figs[i].show()
                figs[i].canvas.flush_events()
            
            input("Press Enter") 
            
    return restart_needed

# DONE store in Model_{idx} class_center_F0
# DONE  in Model_{idx} change width of column B: (based on rows starting from 6th)
# DONE in Model_{idx} add std_relative / mean_relative vs F0_ref_relative_reference == 100Hz
# TODO ??? use formula in Excel with possibility to change F0_ref_relative_reference << use class_center_F0 in formalae for std_relative
# DONE in "Models" add best std_relative
# TODO add in Model_{idx} max(accur), min(abs(mean)), min(std). min(abs(mean_relative)), min(std_relative)
# TODO add in Model_{idx} variance of (accur/std/mean per class)
# TODO add in Model_{idx} outliers per class ??? (+variance)
# DONE mark best results per class with green bold and worst with red bold (might use conditional formating)
# TODO add figures global and per model
#   - train accu / infer accu
#   - mean / mean_rel per class
#   - std / std_rel per class
#   - outliers per class
#   + average values
# TODO mark with bold green/red best and worst parameters in "Models"

def main_processing(datetime_str, main_folder, validation_signals, validation_signals_tag_str, use_date_time_str, do_split_xlsx_by_classes_mode, do_split_by_validation_signals, do_split_by_model, restart_datetime_threshold_str=None):
    for validation_signal_folder in validation_signals:
        # # # main_trained_models_folders = {
        # # #     "inf_cfg_2022_02_15_test_SNR_C_351/f1nw0000",
        # # #     "inf_cfg_2022_02_15_test_SNR_C_351/f2nw0000",
        # # #     "inf_cfg_2022_02_15_test_SNR_C_351/f3nw0000",
        # # #     "inf_cfg_2022_02_15_test_SNR_C_351/f4nw0000",
        # # #     "inf_cfg_2022_02_15_test_SNR_C_351/f5nw0000",
        # # #     "inf_cfg_2022_02_15_test_SNR_C_351/m1nw0000",
        # # #     "inf_cfg_2022_02_15_test_SNR_C_351/m2nw0000",
        # # #     "inf_cfg_2022_02_15_test_SNR_C_351/m3nw0000",
        # # #     "inf_cfg_2022_02_15_test_SNR_C_351/m4nw0000",
        # # #     "inf_cfg_2022_02_15_test_SNR_C_351/m5nw0000",
        # # # }
        # # main_trained_models_folders = {f"inf_cfg_2022_02_15_test_SNR_C_351/{validation_signal_folder}"}
        # # models_filename_tag_str = f"inf_cfg_2022_02_15_test_SNR_C_351_{validation_signal_folder}"
        # main_trained_models_folders = {f"inf_cfg_2022_02_15_test_SNR_C_100/{validation_signal_folder}"}
        # models_filename_tag_str = f"inf_cfg_2022_02_15_test_SNR_C_100_{validation_signal_folder}"
        # do_split_xlsx_by_classes_mode = True

        # main_trained_models_folders = {f"SNR_test_2022_02_20/{validation_signal_folder}"}
        # models_filename_tag_str = f"SNR_test_2022_02_20_{validation_signal_folder}"
        # do_split_xlsx_by_classes_mode = True
        # max_epoch = 100

        # TODO iterate through different validation data SNR sets
        # SNR_dB_idx = validation_signal_folder.find("{SNR_dB}")
        # if SNR_dB_idx > -1:
        #     TODO 1. odczytać jakie są dostępne katalogi pasujące do wzorca
        #     TODO 2. przetworzyć kolejne katalogi stosująć ten sam models_filename_tag_str (ten sam plik xlsx)
        #     TODO 2a. sprawdzić czy różne sygnały (różne SNRy) tworzą różne modele (?) // może lepiej jednak łączyć w ramach jednego modelu? <= raczej na razie nie
        #     TODO 2b. różne SNR_dB - ten sam model: te same train_accur ale różne infer_accur 
        #     validation_signal_folder_SNR  = validation_signal_folder
        #     validation_signal_folder_SNR[SNR_dB_idx:(SNR_dB_idx+8)] = ???

        main_trained_models_folders = {f"{main_folder}/{validation_signal_folder}"}

        # ======================================
        # main_trained_models_folders = {
        #     "lr_test_2022_02_15_m1nw0000",
        # }
        # models_filename_tag_str = "lr_test_2022_02_15_m1nw0000"
        # # do_split_xlsx_by_classes_mode = False
        # do_split_xlsx_by_classes_mode = True

        # main_trained_models_folders = {
        #     "lr_test_2022_02_15_f1nw0000",
        # }
        # models_filename_tag_str = "lr_test_2022_02_15_f1nw0000"
        # do_split_xlsx_by_classes_mode = True

        # main_trained_models_folders = {
        #     "lr_test_2022_02_15_f5nw0000",
        # }
        # models_filename_tag_str = "lr_test_2022_02_15_f5nw0000"
        # do_split_xlsx_by_classes_mode = True

        # ======================================
        # main_trained_models_folders = {
        #     "lr_test_2022_02_03_synth",
        # }
        # models_filename_tag_str = "lr_test_2022_02_03_synth"
        # do_split_xlsx_by_classes_mode = True

        # ======================================
        xls_store_per_class_data = False
        # ======================================

        process_folder_cfg = {
            "do_split_xlsx_by_classes_mode": do_split_xlsx_by_classes_mode,
            "xls_store_per_class_data": xls_store_per_class_data,
            "do_split_by_validation_signals": do_split_by_validation_signals,
            "do_split_by_model": do_split_by_model, 
            "validation_signal_folder": validation_signal_folder,
            "use_date_time_str": use_date_time_str,
            "main_folder": main_folder,
            "validation_signals_tag_str": validation_signals_tag_str, 
            "datetime_str": datetime_str,

            "do_pause": False,
        }

        for main_trained_models_folder in main_trained_models_folders:
            print(f"Processing: {main_trained_models_folder}")

            SNR_dB_flag_idx =  Path(main_trained_models_folder).name.find(r"{SNR_dB}")

            found_folders = list()
            if SNR_dB_flag_idx > -1:
                # find filters fitting the pattern
                folder_mask_head = Path(main_trained_models_folder).name[:SNR_dB_flag_idx+1]
                folder_mask_tail = Path(main_trained_models_folder).name[SNR_dB_flag_idx+8-1:] # leave '}'

                folder_mask = folder_mask_head + '*' 
                potential_folders = [file.name for file in Path("./output_data/inference", main_trained_models_folder).parent.glob(folder_mask)]

                for folder_name in potential_folders:
                    folder_candicate = Path(Path(main_trained_models_folder).parent, folder_name)
                    # check if this is a folder
                    if Path("./output_data/inference", folder_candicate).is_dir():
                        # check if we can read SNR
                        tmp_str = folder_candicate.name[SNR_dB_flag_idx+1:]
                        validation_SNR_str = ""
                        # check for sign
                        if tmp_str[0] == '-':
                            validation_SNR_str += '-'
                            tmp_str = tmp_str[1:]
                        
                        while (tmp_str[0] >= '0') and (tmp_str[0] <= '9'):
                            validation_SNR_str += tmp_str[0]
                            tmp_str = tmp_str[1:]
                        if tmp_str[0] == '.':
                            validation_SNR_str += '.'
                            tmp_str = tmp_str[1:]
                        else:
                            # wrong format
                            continue
                        while (tmp_str[0] >= '0') and (tmp_str[0] <= '9'):
                            validation_SNR_str += tmp_str[0]
                            tmp_str = tmp_str[1:]

                        # TODO check the number in SNR_str
                        print(validation_SNR_str) 

                        # check if the mask tail fits the pattern
                        if folder_mask_tail != tmp_str:
                            # wrong format
                            continue
                        
                        # found_folders.append(Path(Path(main_trained_models_folder).parent, folder_candicate.name))
                        found_folders.append((folder_candicate, folder_mask_head + validation_SNR_str + folder_mask_tail, validation_SNR_str))
            else:
                found_folders.append((main_trained_models_folder, process_folder_cfg["validation_signal_folder"], None))

            for found_main_trained_models_folder, validation_signal_folder, validation_SNR_str in found_folders:
                process_folder_cfg_tmp = process_folder_cfg.copy()
                process_folder_cfg_tmp["validation_signal_folder"] = validation_signal_folder
                while process_folder(found_main_trained_models_folder, process_folder_cfg_tmp, validation_SNR_str, restart_datetime_threshold_str) == True:
                    print("Some subfolders have to be processed again. Restarting ....")

def main():

    main_trained_models_folders = {
        # "trainer_test_2022_01_11_A_synth_40dB",
        # "trainer_test_2022_01_11_B_synth_40dB",
        # "trainer_test_2022_01_11_C_synth_40dB",
        # "trainer_test_2022_01_11_C_100_synth_40dB",
        # "trainer_test_2022_01_23_D_synth_40dB",
        #
        "lr_test_2022_02_03_synth",
        "lr_test_2022_02_03_f1nw0000",
    }
    models_filename_tag_str = "lr_test_2022_02_03"
    do_split_xlsx_by_classes_mode = False

    do_split_by_model = False

    # ============================================== #
    # main_folder = "SNR_0-50_misc_act_test_2022_02_22"
    # do_split_by_validation_signals = True # if False results for different validation signals are stored in the same xlsx file

    # main_folder = "SNR_0-50_misc_act_test_2022_03_02"
    # do_split_by_validation_signals = False # if False results for different validation signals are stored in the same xlsx file

    main_folder = "SNR_0-50_misc_act_test_2022_03_05b"
    do_split_by_validation_signals = False # if False results for different validation signals are stored in the same xlsx file

    main_folder = "SNR_0-50_misc_act_test_2022_03_05b2"
    do_split_by_validation_signals = False # if False results for different validation signals are stored in the same xlsx file

    # main_folder = "SNR_0-50_misc_act_test_2022_03_11"
    # do_split_by_validation_signals = False # if False results for different validation signals are stored in the same xlsx file

    # main_folder = "SNR_10_misc_act_test_2022_02_22"
    main_folder = "SNR_10_misc_act_test_2022_02_22_N_351_log"
    do_split_by_validation_signals = False # if False results for different validation signals are stored in the same xlsx file



    main_folder = "const_SNR_test_2022_03_23_Conv"
    # do_split_by_validation_signals = True # if False results for different validation signals are stored in the same xlsx file
    # do_split_by_model = False
    do_split_by_validation_signals = False # if False results for different validation signals are stored in the same xlsx file
    do_split_by_model = True

    main_folder = "const_SNR_test_2022_03_23_MLP2"
    # do_split_by_validation_signals = True # if False results for different validation signals are stored in the same xlsx file
    # do_split_by_model = False
    do_split_by_validation_signals = False # if False results for different validation signals are stored in the same xlsx file
    do_split_by_model = True

    
    # # for KSIiT'2022
    # main_folder = "trainer_test_2022_01_23_D_synth_40dB"
    # main_folder = "KSTiT_reload_test_2022_04_20_MLP2_seg_5000"
    # main_folder = "KSTiT_reload_test_2022_04_20_MLP2_seg_10000"
    # main_folder = "KSTiT_reload_test_2022_04_20_MLP2_seg_20000"
    main_folder = "KSTiT_reload_test_2022_04_20_MLP2_seg_25000"
    # main_folder = "KSTiT_reload_test_2022_04_20_MLP2_seg_50000"
    # main_folder = "KSTiT_reload_test_2022_04_20_Conv"
    # do_split_by_validation_signals = True # if False results for different validation signals are stored in the same xlsx file
    # do_split_by_model = False
    do_split_by_validation_signals = True # if False results for different validation signals are stored in the same xlsx file
    do_split_by_model = False

    # main_folder = "const_SNR_test_2022_05_20_Conv"
    main_folder = "const_SNR_test_MLP2_100"
    do_split_by_validation_signals_synth = True # if False results for different validation signals are stored in the same xlsx file
    do_split_by_validation_signals_speech = False # if False results for different validation signals are stored in the same xlsx file
    do_split_by_model = False
    # do_split_by_model = True # for speech signals

    # do_split_by_validation_signals_speech = True # for CMPO smoothing
    # do_split_by_model = False # for CMPO smoothing
    
    restart_datetime_threshold_str = "2022-03-11 10:50"
    do_split_xlsx_by_classes_mode = True

    use_date_time_str = False # use datetime string in xlsx file name
    datetime_str = f"{datetime.now().strftime('%Y-%m-%d-%H%M')}"

    speech_f_signals = [
        "f1nw0000",
        "f2nw0000",
        "f3nw0000",
        "f4nw0000",
        "f5nw0000",
    ]
    speech_m_signals = [
        "m1nw0000",
        "m2nw0000",
        "m3nw0000",
        "m4nw0000",
        "m5nw0000",
    ]
    synth_signals = [
        r"synth{SNR_dB}", # {SNR_dB} signals
    ]

    # process separately data for speech and synthetic validation signals 
    main_processing(datetime_str, main_folder, speech_f_signals, "speech_f", use_date_time_str, do_split_xlsx_by_classes_mode, do_split_by_validation_signals_speech, do_split_by_model, restart_datetime_threshold_str)
    main_processing(datetime_str, main_folder, speech_m_signals, "speech_m", use_date_time_str, do_split_xlsx_by_classes_mode, do_split_by_validation_signals_speech, do_split_by_model, restart_datetime_threshold_str)
    main_processing(datetime_str, main_folder, synth_signals, "synth",  use_date_time_str, do_split_xlsx_by_classes_mode, do_split_by_validation_signals_synth, do_split_by_model, restart_datetime_threshold_str)


if __name__ == "__main__":
    # TODO add train_SNR to xlsx
    # TODO speech and synthetic validation data results split into different xlsx files
    main()